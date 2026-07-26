#!/usr/bin/env python3
"""Build the Peptide Founder Media Intelligence workbook.

The workbook intentionally separates interview statements from current verification.
It excludes dosing and personal-use instructions and is not medical or legal advice.
"""

from __future__ import annotations

from collections import Counter
from datetime import date
import json
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


AS_OF = "2026-07-26"
OUTPUT = Path(__file__).with_name("Peptide-Founder-Media-Intelligence.xlsx")
WEB_OUTPUT = Path(__file__).with_name("peptide-founder-media-intelligence-data.js")

INK = "26211D"
ESPRESSO = "382E28"
CREAM = "F7F2E9"
PORCELAIN = "FCFAF6"
STONE = "E8E0D4"
MUTED = "746A62"
SAGE = "57705B"
SAGE_LIGHT = "DFE9DF"
COBALT = "305EA8"
COBALT_LIGHT = "DEE8F7"
GOLD = "B59655"
GOLD_LIGHT = "F4E9CE"
CLAY = "A66059"
CLAY_LIGHT = "F1DEDB"
RED = "A33E3E"
RED_LIGHT = "F6DDDB"
WHITE = "FFFFFF"

THIN_STONE = Side(style="thin", color=STONE)
BORDER = Border(left=THIN_STONE, right=THIN_STONE, top=THIN_STONE, bottom=THIN_STONE)

EVIDENCE_LABELS = {
    "A": "A - Public record / regulator / court / filing",
    "B": "B - Independent media or directory",
    "C": "C - Company-controlled source",
    "D": "D - Founder, guest, show, or agency claim",
    "E": "E - Review, complaint, or litigant allegation",
    "I": "I - Analyst inference",
}


def source(
    source_id: str,
    priority: str,
    lane: str,
    company: str,
    operator: str,
    title: str,
    published: str,
    duration: str,
    stage: str,
    model: str,
    media_url: str,
    current_url: str,
    current_status: str,
    freshness: str,
    confidence: str,
    listen_for: str,
) -> dict:
    return {
        "ID": source_id,
        "Priority": priority,
        "Lane": lane,
        "Company": company,
        "Founder / operator": operator,
        "Episode / source": title,
        "Published": published,
        "Duration": duration,
        "Stage at recording": stage,
        "Model": model,
        "Media link": media_url,
        "Current-check link": current_url,
        "Current status at 2026-07-26": current_status,
        "Freshness": freshness,
        "Confidence": confidence,
        "Why it matters": listen_for,
    }


SOURCES = [
    source(
        "JOI",
        "1 - Start here",
        "Exact category founder",
        "Joi + Blokes",
        "Josh and Katy Whalen",
        "The New Health Gold Rush: Hormones, Peptides, GLP-1s & Fertility",
        "2026-05-14",
        "59:45",
        "Founded 2021; about five years old",
        "Self-pay diagnostics-first telehealth; separate female and male brands share operations",
        "https://www.youtube.com/watch?v=aigF8IfN4Is",
        "https://www.entrepreneur.com/building-a-business/they-built-a-50-million-business-before-glp-1s-became-mainstream-now-theyre-eyeing-this-new-health-frontier",
        "Active; acquired HerMD and announced fertility expansion. Private scale claims are not audited.",
        "Current; refreshed",
        "4/5 model, 2/5 financials",
        "Best direct evidence for a female-leading front, male sibling brand, shared backend, and category-education moat.",
    ),
    source(
        "PEPTIRA",
        "1 - Start here",
        "Exact category founder",
        "Peptira + Peptide Price",
        "Benjamin Criss and Derek Pruski",
        "How We Got Started: The Story of Peptide Price + Peptira",
        "2025-11-01",
        "46:59",
        "Peptira about five months old",
        "RUO ecommerce plus comparison, referral, education, and paid community",
        "https://www.youtube.com/watch?v=Be58yiS6Vs4",
        "https://peptira.com/test-results/",
        "Both active. Peptira expanded its catalog and Peptide Price community; current shipping remedies are stricter than the episode promise.",
        "Recent but materially changed",
        "3/5 overall",
        "Most useful day-one RUO operating interview: sourcing sequence, labels, lot IDs, COAs, support, pack ladders, and community feedback.",
    ),
    source(
        "EC325",
        "1 - Start here",
        "Fast-growth case",
        "Anonymous US peptide brand",
        "Budai Media",
        "Ecom Show #325: If I Had to Scale a Peptide Brand Today",
        "2026-07-21",
        "15:15",
        "Claimed six months after a November 2025 start",
        "Anonymous WooCommerce RUO brand; agency-operated acquisition and lifecycle",
        "https://shows.acast.com/theecomshow/episodes/325-budai-media-specials-if-i-had-to-scale-a-peptide-brand-t",
        "https://www.fda.gov/drugs/drug-alerts-and-statements/fdas-concerns-unapproved-glp-1-drugs-used-weight-loss",
        "Freshest operating update. Agency calls the model a near-term window and says Google scaled more reliably than Meta.",
        "Very current; anonymous",
        "2/5 revenue, 4/5 agency experience",
        "Freshest exact growth playbook, but treat every revenue and margin number as an unaudited agency claim.",
    ),
    source(
        "PERELEL",
        "1 - Start here",
        "Female-leading analog",
        "Perelel",
        "Alex Taylor, Victoria Thain Gioia, Dr. Banafsheh Bayati",
        "Success Almost Broke Her: Building a Multimillion-Dollar Wellness Brand",
        "2026-01-04",
        "35:16",
        "Five years old; just under 40 employees",
        "Doctor-founded DTC subscription brand organized around women's life stages",
        "https://podcasts.apple.com/us/podcast/success-almost-broke-her-this-recovering-girlboss-built/id1593458926?i=1000743844412",
        "https://www.prnewswire.com/news-releases/perelel-secures-27m-growth-investment-from-prelude-growth-partners-to-reimagine-womens-health-302622022.html",
        "Active and expanded; $27M growth investment confirmed. Female-leading master brand still carries one men's product.",
        "Current; high fit",
        "5/5 tactics, 3-4/5 private metrics",
        "Strongest brand-architecture and creator-cohort analog for turning interchangeable products into an identity system.",
    ),
    source(
        "SHEMED",
        "1 - Start here",
        "Female-leading analog",
        "SheMed",
        "Olivia and Chloe Ferro",
        "The Weightloss Wonders",
        "2025-04-24",
        "31:23",
        "About four months after launch",
        "Clinician-led UK telehealth subscription for women, primarily ages 35 to 55",
        "https://www.buzzsprout.com/850540/episodes/17034792-the-weightloss-wonders",
        "https://www.shemed.co.uk/about",
        "Company says 100K+ women served; reported $50M raise. Male sibling HeMed launched on shared infrastructure.",
        "Historical launch story; current state refreshed",
        "5/5 model, 4/5 private scale",
        "Clearest female-front/shared-backend precedent and best example of testing brand tone against the real buyer.",
    ),
    source(
        "MOCHI",
        "1 - Start here",
        "Platform analog",
        "Mochi Health",
        "Dr. Myra Ahmad",
        "Inside the Peptide Craze",
        "2026-04-28",
        "1:08:00; founder segment 43:41-68:19",
        "Four years old; minimally funded marketplace",
        "Recurring cash-pay access connecting customers, providers, pharmacies, testing, and support",
        "https://podcasts.apple.com/us/podcast/inside-the-peptide-craze-myra-ahmad-ceo-of-mochi/id1774403992?i=1000764115940",
        "https://joinmochi.com/longevity",
        "Active; launched a longevity offering after the interview. Company claims 500K+ members; provider count varies by source.",
        "Very current; rapidly changing",
        "4/5 model, 2-3/5 private metrics",
        "Best analog for recurring access, price transparency, provider continuity, testing capacity, and restrained funding.",
    ),
    source(
        "VITABELLA",
        "2 - Deep dive",
        "Exact category founder",
        "Vita Bella Health",
        "Phil Vella",
        "The State of the Hormone & Peptide Industry",
        "2026-02-26",
        "1:14:19",
        "Founded 2023; about three years old",
        "Cash-pay telehealth MSO and recurring membership",
        "https://www.youtube.com/watch?v=vwgd2UFdE6A",
        "https://vitabella.com/membership/",
        "Active; California added after recording. Current pricing is visible, while state, review, and legacy-price claims conflict across pages.",
        "Recent; partly changed",
        "3/5 overall",
        "Useful for membership mechanics, service capacity, partner boundaries, and why a broad catalog is operationally hard.",
    ),
    source(
        "REDLINE",
        "2 - Deep dive",
        "Day-zero operator",
        "Redline Biolabs",
        "Kyle and Ryan",
        "Two Firefighters Built a Peptide Company",
        "2026-05-26",
        "1:19:22",
        "Company domain created 2026-04-04",
        "Founder-led RUO ecommerce",
        "https://www.youtube.com/watch?v=JQSJYACbOCU",
        "https://www.redlinebiolabs.com/lab-reports",
        "Active store and lab-report library; no evidence of rapid scale. Current site contains documentation and affiliate-language inconsistencies.",
        "Very recent; pre-scale",
        "4/5 footprint, 1/5 scale/quality claims",
        "Best day-zero case for founder visibility, payment failure, supplier loss, overselling, service recovery, and channel codes.",
    ),
    source(
        "EC320",
        "2 - Deep dive",
        "Fast-growth case",
        "Anonymous US peptide brand",
        "Budai Media",
        "Ecom Show #320: How We Scaled a US Peptide Brand",
        "2026-05-12",
        "17:48",
        "Claimed five months after launch",
        "Anonymous WooCommerce RUO brand",
        "https://shows.acast.com/theecomshow/episodes/320-budai-media-specials-how-we-scaled-a-us-peptide-brand-fr",
        "https://shows.acast.com/theecomshow/episodes/325-budai-media-specials-if-i-had-to-scale-a-peptide-brand-t",
        "Superseded by #324 and #325. Initial $500K monthly claim later became $600K.",
        "Recent but superseded",
        "2/5",
        "Original channel ramp, learning-period, email-provider failure, and fulfillment/payment coordination story.",
    ),
    source(
        "EC324",
        "2 - Deep dive",
        "Fast-growth case",
        "Anonymous US peptide brand",
        "Budai Media",
        "Ecom Show #324: How to Scale Your Peptide Brand",
        "2026-07-07",
        "14:35",
        "Claimed five months after launch",
        "Anonymous WooCommerce RUO brand",
        "https://shows.acast.com/theecomshow/episodes/324-budai-media-specials-how-to-scale-your-peptide-brand-so",
        "https://shows.acast.com/theecomshow/episodes/325-budai-media-specials-if-i-had-to-scale-a-peptide-brand-t",
        "Mostly current, but #325 is the latest agency position. Approval-bypass discussion is excluded from the workbook playbook.",
        "Very recent; superseded in part",
        "2/5 revenue, 4/5 agency experience",
        "Lifecycle flows, same-product volume ladders, subscriptions, and channel-fragility observations.",
    ),
    source(
        "REM",
        "2 - Deep dive",
        "Young adjacent operator",
        "Remedora",
        "Coy McAlexander",
        "Advanced Affiliate EP 134",
        "2026-05-28",
        "55:25",
        "Very young; domain created 2025-10-23",
        "Telehealth launch and operating layer",
        "https://www.youtube.com/watch?v=1qgxkaPcq4k",
        "https://remedora.com/",
        "Active $200/month platform claim. Scale and vertical-integration assertions remain unverified; retatrutide-compounding discussion is outdated.",
        "Recent but one material claim is outdated",
        "3/5 footprint, 1-2/5 scale",
        "Useful for networking, platformization, cohort LTV, and concentration risk. Not a legal or compounding authority.",
    ),
    source(
        "ARRAE",
        "2 - Deep dive",
        "Creator-growth analog",
        "Arrae",
        "Siffat Haider and Nish Samantray",
        "What Makes a Successful Wellness Brand",
        "2023-05-02",
        "39:21",
        "Three years old; about 12-13 employees",
        "Creator-native DTC wellness and supplement brand",
        "https://podcasts.apple.com/us/podcast/siffat-haider-co-founder-co-ceo-of-arrae-what-makes/id1432195522?i=1000611429597",
        "https://www.forbes.com/sites/karineldor/2026/02/18/building-100m-wellness-brand-arrae-siff-haider-on-calculated-risk-and-relentless-standards/",
        "Active with broader assortment, subscription, and major retail. Company reports $100M+ cumulative revenue and 1.6M+ units, unaudited.",
        "Old launch tactics; current state refreshed",
        "5/5 tactics, 3/5 metrics",
        "Strongest creator-seeding system: broad honest gifting, promote organic advocates, multi-year relationships, and customer-as-creator.",
    ),
    source(
        "BELLI",
        "2 - Deep dive",
        "Community-growth analog",
        "BelliWelli",
        "Katie Wilson, Nick Wilson, Tyson",
        "How Katie Wilson Bought Facebook Groups to Build a Gut Health Empire",
        "2025-06-27",
        "57:17",
        "Post-DTC launch; scaling into Walmart and Target",
        "DTC subscription plus mass retail wellness",
        "https://podcasts.apple.com/us/podcast/former-matchmaker-turned-ceo-how-katie-wilson-bought/id1484646914?i=1000714873422",
        "https://www.uschamber.com/co/good-company/launch-pad/health-startups-community-retail-growth",
        "Active powder and gummy assortment, DTC subscription, broad retail, and independently reported $15.4M Series A.",
        "Current; high creator/content fit",
        "5/5 tactics, 3/5 metrics",
        "Best community/content engine: ownable repeatable format, constant field research, and the danger of confusing niche PMF with mass PMF.",
    ),
    source(
        "TRANSCEND",
        "3 - Historical caution",
        "Exact category founder",
        "Transcend Company",
        "Ernest Colling",
        "What's Really Going On in the Peptide Business",
        "2024-09-16",
        "1:01:54",
        "Almost four years old",
        "Telehealth MSO supporting independent practices and pharmacies",
        "https://www.youtube.com/watch?v=c6jUD2kyyDA",
        "https://transcendcompany.com/pages/change-of-provider-letter",
        "Operating under new owner Lance Converse and changed provider structure after a dispute and litigation. Use only for historical origin and scale-break lessons.",
        "Materially changed",
        "3/5 current facts, 2/5 private claims",
        "A cautionary scaling case: creator demand outran support and partner capacity; old founder titles no longer describe current control.",
    ),
    source(
        "BIOLONG",
        "3 - Historical caution",
        "Exact category founder",
        "BioLongevity Labs",
        "Jay Campbell",
        "Mind Pump 2500: Weird Performance-Enhancing Supplements",
        "2024-12-30",
        "Business segment near 1:05:55",
        "About two months old by founder account",
        "RUO storefront plus separate supplements, founder media, courses, and private community",
        "https://mindpumppodcast.com/2500-weird-performance-enhancing-supplements-that-work-with-jay-campbell/",
        "https://biolongevitylabs.com/research/",
        "Active with current content and catalog. Site product-count claims conflict; proposed pen products were not found.",
        "Historical interview; active business refreshed",
        "3/5 footprint, 1-2/5 scale",
        "Useful for education-led demand and existing-community leverage; weak proof of scale or delivered roadmap.",
    ),
]


NOTES: list[dict] = []


def note(
    source_id: str,
    timecode: str,
    topic: str,
    extraction: str,
    application: str,
    evidence: str = "D",
    freshness: str = "From episode; verify before acting",
) -> None:
    start_seconds = None
    if timecode and timecode[0].isdigit():
        first = timecode.split("-")[0].strip()
        parts = first.split(":")
        try:
            if len(parts) == 2:
                start_seconds = int(parts[0]) * 60 + int(parts[1])
            elif len(parts) == 3:
                start_seconds = int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        except ValueError:
            start_seconds = None
    base_url = next(item["Media link"] for item in SOURCES if item["ID"] == source_id)
    timestamp_url = base_url
    if start_seconds is not None and "youtube.com/watch" in base_url:
        joiner = "&" if "?" in base_url else "?"
        timestamp_url = f"{base_url}{joiner}t={start_seconds}s"
    NOTES.append(
        {
            "Source ID": source_id,
            "Company": next(item["Company"] for item in SOURCES if item["ID"] == source_id),
            "Timestamp": timecode,
            "Topic": topic,
            "What the operator said / what happened": extraction,
            "What Biologix can learn": application,
            "Evidence class": EVIDENCE_LABELS[evidence],
            "Freshness note": freshness,
            "Timestamp link": timestamp_url,
        }
    )


# Joi + Blokes
note("JOI", "01:30-03:35", "Founder-market fit", "The founders trace the company to their own difficult search for better care and say they personally believe in the offers.", "Use authentic founder involvement only where it is true; founder conviction can humanize an opaque category.")
note("JOI", "03:38-04:18", "Category education", "Early acquisition required teaching the market about proactive labs and peptides. Later competition reduced the basic education burden.", "Move content from category definition toward operator-specific trust, proof, and decision support as awareness rises.")
note("JOI", "04:19-09:22", "Operating barrier", "State rules, pharmacy shipping, and clinical oversight were major launch barriers. They position themselves between conventional medicine and unregulated RUO sellers.", "Do not borrow clinical positioning without clinical infrastructure; decide and communicate the actual operating lane.")
note("JOI", "08:04-11:54", "Category criticism", "They criticize research-product sellers that provide access without clinical guidance.", "Expect credible operators to attack the category's trust gap. Documentation and role clarity must answer the objection without pretending to provide care.")
note("JOI", "14:40-17:47", "Product concentration", "They say regulatory change reduced the peptide range after 2023 and that they intentionally stayed conservative.", "Build scenario plans for losing a hero category rather than treating the catalog as permanent.")
note("JOI", "18:02-20:24", "Revenue concentration", "Founder says GLP-1 reached about half of revenue before manufacturer pressure caused them to remove public advertising and listings; he says it is now 15-20%.", "Track share of revenue by product and by acquisition promise. Concentration can hide inside a broad-looking catalog.", "D", "Private revenue mix; unaudited")
note("JOI", "29:40-34:05", "Founder strain", "They describe 2025 as their hardest year as CAC rose and leadership stress grew.", "Model leadership and service capacity as constraints, not just ad spend and inventory.")
note("JOI", "31:51-32:20", "Creative ownership", "Josh says he personally studies advertisements and website user journeys.", "Keep a founder or senior owner close to creative quality and conversion, even after hiring agencies.")
note("JOI", "34:06-37:51", "Female-front architecture", "The business runs two audience-facing companies on one operation. Josh calls Joi the stronger engine and says women often bring male partners into Blokes.", "Strong direct support for a female-leading flagship, technical male access, and shared backend.")
note("JOI", "48:53-52:17", "Scale and capital", "They say they are approaching $50M in revenue, are essentially self-funded, and now need more experienced operators or capital.", "Demand can outrun the founders' operating altitude. Define when to add experienced functional leaders.", "D", "Revenue is management-reported")
note("JOI", "52:42-54:36", "Adjacency", "The audience is becoming younger and fertility is presented as the next major adjacency.", "Expand only after the core engine is stable; use customer progression to identify the next front door.")
note("JOI", "55:24-57:31", "Growth versus profitability", "They describe tension between maintaining profitability and preserving venture-attractive growth.", "Choose the capital model deliberately; do not let fundraising optics dictate channel spend.")

# Peptira + Peptide Price
note("PEPTIRA", "00:00-03:12", "Comparison-layer origin", "Peptide Price began as a two-day MVP listing four vendors because price and vendor comparison was difficult.", "A decision tool can become distribution infrastructure before it becomes polished software.")
note("PEPTIRA", "02:41-03:08", "Vendor relationship", "Derek says vendors opt in and his owner relationships let him intervene when customers have problems.", "Partner channels are strongest when they can resolve service failures, not merely drop affiliate links.")
note("PEPTIRA", "03:15-04:31", "Launch channel", "TikTok launched the comparison tool; a viral post allegedly reached about 70K views while he manually sent a raw URL.", "Manual distribution is acceptable before automation. Validate message-to-click behavior first.", "D", "View count is founder-reported")
note("PEPTIRA", "04:32-06:35", "Community-led assortment", "Community requests determined vendors and products; testing and COAs became the main qualification standard.", "Use support tickets, search logs, creator comments, and comparison queries as a structured catalog research feed.")
note("PEPTIRA", "06:37-08:58", "Platform fragility", "Derek says eight TikTok accounts were banned and he moved the audience into a paid Skool community.", "Always build an owned audience and exportable identity graph; rented reach can disappear.")
note("PEPTIRA", "09:00-13:41", "Decision criteria", "Comparison value extends beyond lowest price to per-mg price, codes, test scope, shipping, and reviews.", "Win the comparison surface with structured proof, not generic brand language.")
note("PEPTIRA", "14:38-17:27", "Launch sequence", "Peptira formed in June 2025 and spent roughly six weeks on supply, labels, and packaging before launch.", "Treat sourcing, label system, testing, and fulfillment as the launch critical path.")
note("PEPTIRA", "17:31-21:19", "Batch trust", "Labels evolved from QR-only to include batch ID, measured content, and purity; QR links to batch COAs.", "Put lot identity and meaningful test outputs on the physical product and PDP, with current and historical status.")
note("PEPTIRA", "21:21-22:45", "Volume ladder", "The store uses one, three, five, and ten-unit quantities with volume discounts.", "Test simple same-product quantity ladders before complex mixed bundles.")
note("PEPTIRA", "22:51-24:15", "Unsupported scale claim", "Ben calls Peptira a top-15 US peptide company and promises prices will not increase.", "Never use rankings or permanent-price promises without an auditable basis.", "D", "Unsupported")
note("PEPTIRA", "24:20-29:24", "Service moat", "Ben claims under-50-minute support, GPS on nearly all shipments, and automatic lost-package replacement; HubSpot was planned.", "Instrument first-response time, resolution time, replacement cost, and shipment exception rate before making promises.", "D", "Replacement promise is now contradicted by live policy")
note("PEPTIRA", "30:01-31:38", "Release gate", "They withheld a promotion-ready product until third-party testing returned.", "Make test clearance a hard inventory-release state, not a marketing preference.")
note("PEPTIRA", "31:40-39:49", "Founder overload", "Ben left insurance work and was still packing and handling support late at night, then recognized the need for delegation.", "Document SOPs and staff the exception queue before volume turns founder responsiveness into a bottleneck.")
note("PEPTIRA", "39:52-46:18", "Education risk", "Podcast topics and courses were community-led, but some content crosses toward personal-use protocols.", "Keep education focused on lawful research, testing, documentation, and product handling boundaries; exclude dosing and personal-use guidance.", "D", "Risk remains current")

# Transcend
note("TRANSCEND", "02:24-18:13", "Origin", "Colling links the company to personal health problems, a closed local clinic, and a late-2020 launch.", "Personal frustration can identify demand, but it does not replace the operating and evidence model.")
note("TRANSCEND", "20:03-22:50", "Concierge service", "He kept his day job for about a year while building an education-heavy concierge model.", "Preserve runway while validating high-touch service; do not overhire before demand is repeatable.")
note("TRANSCEND", "24:05-26:34", "Hidden complexity", "State licensing, malpractice, physician-patient rules, and pharmacy shipping were larger obstacles than expected.", "Map every partner, state, and handoff before projecting national scale.")
note("TRANSCEND", "27:20-28:44", "Initial audience", "Initial acquisition focused on bodybuilders and athletes through education and influencers.", "A sharp founding niche can create density, but plan how the brand broadens without losing credibility.")
note("TRANSCEND", "35:07-36:21", "Breakout trigger", "FitCon Dallas and two affiliates allegedly created enough demand for the founder to leave his job within about 90 days.", "Creator pilots should be capacity-gated. A successful launch can be more dangerous than a failed one.", "D", "Affiliate contribution unaudited")
note("TRANSCEND", "36:23-37:05", "Bootstrap claim", "The founder says the basement business was profitable from day one and self-funded.", "Low overhead helps, but do not use the claim as a forecast without full unit economics.", "D", "Unaudited")
note("TRANSCEND", "37:06-38:27", "Scale break", "Scaling from hundreds of thousands to millions per month allegedly broke hiring, training, calls, patient capacity, and pharmacy inventory.", "Build a launch-readiness gate across support, inventory, fulfillment, and partner capacity before adding creators.", "D", "Historical; current control changed")
note("TRANSCEND", "38:27-40:14", "Role boundaries", "Independent pharmacies dispense and ship; Transcend manages goals, process design, and administration.", "Map responsibility for every customer-visible failure and write escalation/continuity terms.")
note("TRANSCEND", "40:16-42:34", "Audience expansion", "The business broadened from bodybuilding toward longevity and discussed employer/B2B expansion.", "Expand audience only after the first segment's retention and service model is proven.")
note("TRANSCEND", "42:34-44:19", "Price trend claim", "Colling says peptide prices fell by one-third to two-thirds over five to ten years.", "Assume molecule pricing can commoditize and build value in proof, service, creators, and data.", "D", "No supporting data")
note("TRANSCEND", "46:24-47:05", "Single-product risk", "GLP-1 awareness drove demand, but he did not want one product to define the company.", "Use the hero SKU as an acquisition wedge while monitoring concentration and cross-category movement.")
note("TRANSCEND", "48:39-52:20", "Mission layer", "The foundation focused on veterans and first responders.", "A credible community mission can strengthen identity, but it must be operationally and financially real.")
note("TRANSCEND", "60:10-60:57", "Foundation claim", "Founder claims roughly 100 patients served and $1.5M provided or donated over 18 months.", "Require auditable impact reporting before making mission-based scale claims.", "D", "Tax filings confirm foundation, not amount")

# Vita Bella
note("VITABELLA", "01:08-02:17", "Geographic coverage", "Vella says Vita Bella operated in 40 states and views state complexity as both barrier and moat.", "Treat geographic breadth as an operations matrix, not a marketing line.")
note("VITABELLA", "03:46-05:50", "Founder background", "The idea followed a technology career, severe injury, and frustration with healthcare markups.", "Transfer the service-design and systems mindset, not unsupported pricing accusations.")
note("VITABELLA", "06:43-08:20", "Healthcare as a service", "The founder modeled recurring access on enterprise software and admits starting broad is harder than starting narrow.", "Start with a coherent wedge and only add categories after navigation, support, and economics work.")
note("VITABELLA", "08:21-13:29", "Competitor testing claim", "He says four competitor samples contained unknown or mislabeled contents but shows no reports, names, or chain of custody.", "Never publish competitor-safety accusations without named reports, independent sampling, and defensible custody.", "D", "Unsupported")
note("VITABELLA", "12:35-14:24", "Markup claim", "The founder gives pharmacy-cost and competitor-markup examples without invoices.", "Do not use unaudited unit economics as a pricing anchor.", "D", "Unsupported")
note("VITABELLA", "14:28-21:07", "Rapid growth claim", "He claims hiring 20 providers, 300% growth in weeks, and a 220% increase in new-patient spend after industry disruption.", "A competitor failure can create a demand shock; maintain surge capacity and verify economics before celebrating growth.", "D", "Unaudited")
note("VITABELLA", "21:39-25:45", "Service retention", "Long visits, service, outcomes, and 48-72-hour pharmacy delivery are positioned as retention levers.", "Tie retention to measurable service states, but align marketing promises with governing terms.")
note("VITABELLA", "25:52-29:25", "Jurisdiction choice", "At recording, the company deliberately excluded California and other complex markets.", "Refusing markets can be a strategic capacity decision, not a failure of ambition.", "D", "California has since launched")
note("VITABELLA", "29:27-34:30", "Membership mechanics", "The offer was $129 monthly plus a $99 consultation, separate labs, quarterly visits, one included treatment, and add-ons.", "If using recurring access, separate the fee, included value, product costs, commitment, and cancellation terms unmistakably.")
note("VITABELLA", "43:25-45:07", "Creator incentive design", "Vella criticizes creator incentives that encourage oversized stacks and higher customer spend.", "Pay for qualified, retained customers and compliant content, not larger baskets or product intensity.")
note("VITABELLA", "45:09-50:29", "Future category", "He expects continued demand and more big-pharma entry through patents and new delivery methods.", "Plan for changing consumer expectations and stronger incumbents, not only today's vendor set.")
note("VITABELLA", "57:14-58:47", "Behavioral retention", "App, training, and accountability are presented as retention tools.", "Build retention around useful workflows and support rather than silent auto-renewal.")
note("VITABELLA", "59:02-60:43", "Growth throttle", "Founder says he would pause growth if service quality deteriorated.", "Create explicit stop rules tied to response time, stockouts, payment failures, and complaint rate.")
note("VITABELLA", "66:14-66:20", "Revenue concentration", "Founder says a single category represented about 40% of revenue despite nearly 100 medications.", "Catalog breadth does not equal diversified economics; measure concentration by net contribution.", "D", "Private claim")

# Anonymous Ecom Show growth case
note("EC320", "00:30-01:19", "Claimed growth", "Agency says a basic WooCommerce brand launched in November 2025 and reached roughly $500K trailing-30-day revenue after five months.", "Use as a hypothesis that demand can ramp quickly, never as a forecast.", "D", "Anonymous, unaudited, later updated")
note("EC320", "01:26-02:45", "Prerequisites", "Self-hosted commerce, card processing, fulfillment, and a professional team are described as prerequisites.", "Launch readiness must include stable payments, inventory, fulfillment, and accountable owners.")
note("EC320", "02:51-03:03", "Agency economics", "The agency says it operated Google, Meta, and email on a revenue-share arrangement.", "Evaluate agencies on contribution margin, data ownership, and scope, not attributed revenue alone.", "D", "Agency account")
note("EC320", "04:17-05:51", "Learning period", "Month one allegedly spent $7K on Google for $4.5K attributed revenue; the loss was treated as a 30-60-day learning investment.", "Set a bounded test budget, leading indicators, and kill rules before assuming a loss will mature into scale.", "D", "Anonymous attributed metrics")
note("EC320", "05:51-08:15", "Channel sequence", "Sequence was Shopping, branded search, then Performance Max. Agency claims about $300K attributed revenue on $151K spend.", "Capture existing product-name intent first and segment by product and margin; reconcile platform attribution to settled orders.", "D", "Anonymous attributed metrics")
note("EC320", "09:17-11:43", "Vendor failure", "Email flows were built immediately, but Klaviyo allegedly closed the account and a FluentCRM rebuild took three weeks.", "Export customer and event data continuously and keep a tested migration path.")
note("EC320", "12:01-14:10", "Meta signal", "Meta allegedly produced very high early ROAS at small spend.", "Treat tiny-budget ROAS as a creative signal, not proof of scalable economics.", "D", "Anonymous agency attribution")
note("EC320", "14:22-17:48", "System interaction", "Agency attributes $14K-$15K days to acquisition, email, fulfillment, and payments finally working together.", "Growth is a system state. Instrument failures across the full customer chain.", "D", "Anonymous revenue claim")

note("EC324", "00:30-01:13", "Scale update", "Agency calls the client its fastest, from zero to $20K daily in five months.", "Preserve as an upside case only; require client identity and raw cohort data before benchmarking.", "D", "Anonymous, unaudited")
note("EC324", "01:18-03:18", "Trust surface", "Agency recommends accessible COAs, clear disclaimers, card payments, and avoiding human-use, mixing, dosing, or benefit claims.", "Build proof and operational clarity into the storefront while keeping education inside lawful research boundaries.")
note("EC324", "03:24-04:05", "Scale update", "Claim rises to $600K monthly, including $450K attributed to Google on $150K spend.", "Changing claim definitions are a warning to reconcile gross orders, settled revenue, refunds, and channel attribution.", "D", "Anonymous, unaudited")
note("EC324", "04:05-05:04", "Channel fragility", "Agency says only two of eight client Meta accounts remained active while all eight were active on Google.", "Do not make one ad platform the business. Maintain creator, search, email, and direct channels.", "D", "Agency-reported experience")
note("EC324", "05:11-06:29", "Existing intent", "Product-name search and Shopping allegedly outperform benefit-oriented queries.", "Prioritize clear product taxonomy and high-intent demand capture; avoid unsupported benefit language.")
note("EC324", "06:29-08:31", "Approval evasion", "The episode includes tactics intended to bypass platform review.", "Exclude entirely. Build no workflow around evasion, renamed products, cloaked pages, or misleading review surfaces.", "D", "Intentionally excluded from implementation")
note("EC324", "08:54-10:22", "Owned lifecycle", "Self-hosted email is presented as a response to hosted-provider shutdowns.", "Owning infrastructure reduces vendor concentration but increases deliverability, security, and maintenance responsibility.")
note("EC324", "10:28-12:19", "Lifecycle flows", "Recommended flows are welcome, browse/cart abandonment, post-purchase, VIP, winback, and replenishment; agency claims 5-15% of sales.", "Implement the event model early and measure incremental lift rather than accepting vendor attribution.", "D", "Revenue share is agency-reported")
note("EC324", "12:26-14:16", "Offer architecture", "Same-product volume discounts and subscriptions are favored over mixed-product bundles.", "Test volume ladders first; any subscription must require clear express consent and easy cancellation.")
note("EC324", "14:16-14:35", "Business-model migration", "Agency says its largest client was moving from RUO ecommerce toward telehealth.", "Treat operating-lane optionality as a board-level question, not a cosmetic site change.", "D", "Anonymous client")

note("EC325", "00:30-01:25", "Latest scale claim", "The freshest episode describes $600K in one month after six months.", "Use only as an unverified upside case; do not plug into forecasts.", "D", "Anonymous and unaudited")
note("EC325", "01:31-03:20", "Latest channel view", "Agency now calls Google the strongest acquisition channel; Meta looked attractive but did not scale reliably.", "Test product-intent capture while treating approval and account continuity as uncertain.", "D", "Current agency experience, not a guarantee")
note("EC325", "03:20-04:37", "Campaign structure", "Sequence remains Shopping, search, then Performance Max, segmented by product and margin.", "Organize reporting around SKU contribution and settled order economics.")
note("EC325", "04:37-07:10", "Optimization discipline", "Legitimate advice is to change budgets gradually, change one variable at a time, and review landing pages monthly.", "Adopt experiment logs and one-variable decision windows; discard the episode's approval-evasion material.")
note("EC325", "07:48-11:22", "Email and subscription", "Agency recommends lifecycle and subscription architecture at launch, including preselection.", "Build lifecycle early, but never preselect subscriptions; require clear express consent and visible terms.")
note("EC325", "11:33-12:04", "Campaign threshold", "Regular campaigns begin after the list reaches hundreds or thousands; SMS remains exploratory.", "Start triggered lifecycle first. Add broadcast cadence only when list size and engagement justify it.")
note("EC325", "12:04-14:33", "Payment instability", "Agency again stresses documentation and card payments while acknowledging unstable payment approval.", "Processor redundancy, reserve visibility, chargeback controls, and honest underwriting are launch-critical.")
note("EC325", "14:33-15:15", "Exit-quality warning", "Agency calls RUO ecommerce a near-term window and doubts easy eight- or nine-figure exits.", "Optimize for durable customer data, creator relationships, proof systems, and optionality, not only near-term gross sales.", "D", "Freshest strategic observation")

# Remedora
note("REM", "00:31-04:45", "Prior scale claims", "Coy describes ecommerce, SEO, affiliate experience, and large revenue in earlier insurance businesses, not Remedora.", "Do not transfer a founder's prior-company scale into the new company's traction.", "D", "Prior business only")
note("REM", "06:15-08:01", "Concentration shock", "He says a CMS change collapsed daily ACA call volume from 10K-12K to about 400.", "Model policy and channel concentration as existential risks even when current economics look exceptional.", "D", "Historical founder account")
note("REM", "12:30-13:45", "Networking system", "Coy says he spends two to three hours daily on relationships and used the network to identify telehealth demand.", "Make operator, lab, processor, creator, and partner relationships a scheduled founder channel.")
note("REM", "16:45-17:30", "Market entry", "He entered telehealth through affiliate sales and prior subscription-commerce experience.", "Learn the demand and operating graph before trying to own every layer.")
note("REM", "22:01-23:45", "Platform thesis", "Remedora was conceived as a Shopify-like launch layer for telehealth.", "Shared infrastructure can turn one brand experiment into a portfolio capability, but product claims require verification.")
note("REM", "26:55-30:15", "Vertical integration", "The founder discusses pharmacy acquisition and integration, without proof a transaction occurred.", "Integrate only where control and economics justify capital, licensing, and operational risk.", "D", "Future or exploratory claim")
note("REM", "31:45-35:52", "Cohort economics", "He says clients often lose money in months one and two, making retention and LTV decisive.", "Track contribution by acquisition month through second and third purchases, not first-order ROAS.", "D", "Private cohort claim")
note("REM", "35:52-38:31", "Retatrutide compounding", "The episode discusses possible compounding pathways.", "Discard this section for current planning; FDA now says retatrutide cannot be used in compounding under federal law.", "D", "Outdated and superseded")
note("REM", "40:30-42:45", "Offer", "Episode price was $200 monthly plus certification-related costs.", "Treat platform pricing as only one line item; verify payment, provider, pharmacy, support, and state costs.")
note("REM", "47:30-52:15", "Founder-led sales", "Networking, conferences, access, and credible proof artifacts drive sales.", "Pair relationship selling with auditable artifacts and a repeatable follow-up system.")

# Redline Biolabs
note("REDLINE", "00:45-01:00", "Founding niche", "The company was founded by two Houston-area firefighters.", "A real occupational identity can create trust and a dense initial community.")
note("REDLINE", "03:15-05:00", "Speed to launch", "They describe moving quickly from supplier research to a live website.", "Speed is useful only after supplier, testing, payment, and inventory gates are real.")
note("REDLINE", "06:45-07:30", "Audience wedge", "First responders are the initial niche; founder transparency and accessible education are the differentiation.", "Start with an audience that can recognize and refer the founders.")
note("REDLINE", "14:45-16:32", "Sourcing", "They acknowledge mixed sourcing and express a goal of more US-synthesized supply.", "Publish precise, product-specific sourcing and test scope; aspirations are not present facts.", "D", "Self-reported")
note("REDLINE", "17:30-18:15", "Trust response", "Visible founders and COAs are their answer to the scam objection.", "Put real people beside verifiable batch proof, while making the limits of each proof clear.")
note("REDLINE", "18:15-20:45", "Payment loss", "Early card-processing failures allegedly caused at least six abandoned orders when buyers encountered cryptocurrency.", "Do not launch creator traffic until the primary and backup checkout rails have been tested end to end.", "D", "Founder-reported")
note("REDLINE", "35:45-37:16", "Supplier loss", "They say a supplier completed a small trial order and vanished after a roughly $700 larger order.", "Use staged purchase orders, counterparty verification, retention samples, and acceptance criteria.")
note("REDLINE", "38:16-39:15", "Founder acquisition", "Founder-face TikTok and podcasts are the main early channels.", "In a low-trust category, access and candor can outperform polished anonymous branding.")
note("REDLINE", "40:00-42:15", "Oversell recovery", "An inventory error oversold a bundle; founders called buyers, substituted inventory, and promised the missing item later.", "Synchronize saleable inventory, define substitutions, and proactively own service failures.")
note("REDLINE", "42:30-43:15", "Attribution", "A podcast-specific discount code was created.", "Give every creator and appearance a source ID, code, landing page, and cohort.")
note("REDLINE", "43:45-47:15", "Early service", "Fast replies and proactive stock updates are treated as an early advantage.", "Turn founder responsiveness into documented SLAs and escalation paths before it becomes unsustainable.")

# BioLongevity Labs
note("BIOLONG", "01:05:55-01:06:19", "Ambition", "Jay says the two-month-old business aimed to become the largest research-peptide and bioregulator seller within 12-15 months.", "Separate ambition from evidence. Demand proof is revenue, repeat behavior, and contribution, not a stated goal.", "D", "Historical and unverified")
note("BIOLONG", "01:06:19-01:06:58", "Future format", "He describes future pen-format products from a California facility.", "Do not repeat planned product formats as delivered traction.", "D", "Current catalog check found no pens")
note("BIOLONG", "01:06:58-01:07:27", "Audience leverage", "Founder claims an existing 650-member paid education community and multiple courses.", "An owned, paying knowledge community can lower launch friction, but verify engagement and buyer overlap.", "D", "Unaudited")
note("BIOLONG", "01:07:27 onward", "Education centerpiece", "BioLongevity is described as the commercial centerpiece of a broader education ecosystem.", "Build media and research content as a durable discovery layer, while keeping human-use guidance out of RUO commerce.")

# SheMed
note("SHEMED", "01:33-02:48", "Launch demand", "The female-focused launch expected roughly 1,000 applicants and claimed about 16,000.", "A sharp female front can unlock latent demand, but applicant volume is not the same as paid retention.", "D", "Founder-reported")
note("SHEMED", "04:06-06:37", "Founder insight", "Olivia's long diagnosis journey produced the founding insight.", "Use lived experience to shape questions and service language, not to substitute for expertise.")
note("SHEMED", "08:03-09:24", "Seriousness signal", "Mandatory testing is used to frame the problem as biological, not cosmetic.", "Visible friction can signal seriousness only when it is genuinely necessary, useful, and expertly governed.")
note("SHEMED", "09:34-12:37", "Role clarity", "The founders explicitly are not clinicians. Clinical decisions belong to GPs, nurses, labs, and pharmacies; founders own communication and support.", "Make founder, expert, lab, supplier, and service responsibilities visible instead of blending authority.")
note("SHEMED", "12:37-18:22", "Shareable education", "A first-box QR booklet contains short explainers and shareable resources.", "Turn lot records, test explanations, and research literacy into portable, shareable trust objects.")
note("SHEMED", "18:53-22:28", "Continuity", "Weekly check-ins and monthly reviews create reassurance.", "Design a consistent post-purchase information and support rhythm without drifting into personal-use guidance.")
note("SHEMED", "23:14-24:54", "Channel mix", "Acquisition spans Meta, Google, Reddit, Pinterest, founder recipes, and customer stories.", "Diversify discovery and measure each path separately; do not rely on a single platform.")
note("SHEMED", "25:27-26:22", "Brand testing", "The initial bright visual direction underperformed; conversion testing moved the brand toward muted colors.", "Let the actual buyer's conversion and qualitative response shape tone, not founder taste.")
note("SHEMED", "26:29", "Buyer mismatch", "Most customers were 35-55, older than the founders.", "Research the buyer's life stage, trust needs, language, and visual preferences directly.")
note("SHEMED", "27:22-28:48", "Pre-scale learning", "Research, community conversations, conferences, advisors, mentors, and team quality preceded scale.", "Schedule customer and operator learning before each growth gate.")

# Arrae
note("ARRAE", "05:56-08:40", "Product idea", "The founder wanted something effective, understandable, and visually desirable for a common frustration.", "Commodity products gain meaning through a legible problem and desirable brand world.")
note("ARRAE", "08:40-10:54", "Narrow start", "Arrae began with two common problems rather than a broad catalog.", "Start with enough assortment to prove the category thesis, not enough to maximize SKU count.")
note("ARRAE", "12:06-13:39", "Build order", "Sequence was trusted formulator, workable manufacturer and MOQ, then brand.", "Lock quality and supply constraints before scaling creative promises.")
note("ARRAE", "13:39-15:14", "Focus", "The team asks for the single next action most likely to move the business.", "Run a weekly constraint review and force one primary growth or operations priority.")
note("ARRAE", "15:45-16:21", "Soft launch", "A small community tested pricing, MVP, and messaging before a wider launch.", "Use a controlled creator/customer cohort to validate trust, price, and operations.")
note("ARRAE", "17:27-18:50", "Creator seeding", "With no influencer budget, founders gifted broadly with no posting obligation; one organic post triggered a wider chain.", "Separate honest seeding from paid deliverables and track organic advocates.")
note("ARRAE", "18:51-19:38", "Promote affection", "Early revenue funded paid relationships with creators who had already shown genuine interest; only five or six major year-long partners were maintained.", "Move proven organic advocates into durable paid partnerships rather than renting a large one-off roster.")
note("ARRAE", "19:38-23:25", "Repeated support", "Several creator relationships lasted roughly three years, making recommendations more credible.", "Reward continuity, content quality, and cohort retention, not only last-click revenue.")
note("ARRAE", "23:26-26:10", "Customer-as-creator", "Customers received premium mailers and artifacts comparable to influencers, encouraging sharing.", "Design every fulfilled order as a creator-ready proof and story kit.")
note("ARRAE", "26:16-29:35", "Brand personality", "Ordinary customers are creators; brand social should behave like a person rather than a product grid.", "Build recognizable recurring editorial formats and a human voice.")
note("ARRAE", "27:34-28:05", "Channel ownership", "An early TikTok went viral, but the small team deprioritized it and later had to rebuild.", "Assign a named owner to any channel showing organic pull.")
note("ARRAE", "30:33-33:53", "Capital timing", "The company raised after traction; money went to inventory, people, larger POs, and working capital.", "Use capital to relieve proven constraints, not to manufacture demand.")
note("ARRAE", "34:28-35:50", "Founder division", "Siffat owns brand, community, voice, and email; Nish owns paid media, site optimization, CAC, LTV, AOV, subscription, and finance.", "Pair brand authority with an equally empowered unit-economics operator.")

# Perelel
note("PERELEL", "01:03-01:35", "Taxonomy", "Products follow distinct hormonal and life stages, with medical experts informing formulation and education.", "Organize discovery around understandable needs or research objectives, while keeping claims inside the actual operating lane.")
note("PERELEL", "02:23-10:53", "Complementary founders", "Founder-market fit combines editorial brand experience, personal experience, and expert cofounders.", "Build a leadership triangle across brand, operating economics, and technical authority.")
note("PERELEL", "11:27-13:16", "Commodity differentiation", "Entering a crowded category required a rigorous brand book, distinct values, voice, identity, and visual language.", "Brand must change recognition, trust, and willingness to return, not merely recolor a vial.")
note("PERELEL", "13:17-14:43", "Unique capability", "The forcing question was, 'What can only we do?' Their answer was expert editorial content and community.", "Define a proprietary trust and distribution system competitors cannot copy by swapping packaging.")
note("PERELEL", "14:43-15:49", "Long creator arc", "Creator relationships follow real life-stage journeys; subscriptions advance customers between trimester products.", "Use long narrative arcs and lifecycle states, while never turning an RUO journey into personal-use advice.")
note("PERELEL", "15:51-17:22", "Scrappy launch", "First-year marketing was about $10K; events, seeding, surprise-and-delight, and credibility-bearing partners carried launch.", "Use controlled seeding and high-touch proof before expensive reach.")
note("PERELEL", "17:48-18:30", "Message order", "Early messaging emphasized what the brand was, why it was trustworthy, and how the system worked; paid amplification followed.", "Solve comprehension and trust before buying scale.")
note("PERELEL", "18:31-19:35", "Creator cohorts", "Creator relationships are separated by awareness/reach versus conversion objectives.", "Define KPI, compensation, attribution window, and content role for each creator cohort.")
note("PERELEL", "19:35-20:30", "Word of mouth", "Affiliate, targeted gifting, events, and customer surprise-and-delight reinforce one another.", "Design creator, customer, and community programs as one referral system.")
note("PERELEL", "20:56-21:34", "Capital caution", "Founders warn against excessive capital and unrealistic valuations.", "Preserve strategic freedom until repeat behavior and unit economics are clear.")
note("PERELEL", "21:47-24:01", "Assortment thesis", "Five launch SKUs and 2,500-unit MOQs were needed to prove the life-stage system; they bootstrapped and used a friends-and-family SAFE.", "Every launch SKU must prove the positioning or increase basket logic, not simply make the store look large.")
note("PERELEL", "24:37-25:28", "Operational reach", "Strong content and gifting helped a small team ship to all 50 states within 30 days.", "Content can create demand quickly; fulfillment readiness must precede reach.")
note("PERELEL", "28:40-30:32", "Personal runway", "One founder consulted before launch and for almost a year after to preserve personal runway.", "Protect founder runway so early decisions are not forced by short-term cash pressure.")
note("PERELEL", "32:04-33:39", "Peer network", "Peer founders at comparable stages supplied talent, consultants, and practical support.", "Build a same-stage operator council for current answers, not only advice from giant incumbents.")

# BelliWelli
note("BELLI", "13:10-16:40", "Problem community", "The founder's own experience and conversations with women exposed an underserved gut-health community.", "Use problem communities as listening systems before choosing the brand promise.")
note("BELLI", "17:09-18:41", "Product iteration", "Founder, food scientists, and dietitian iterated for months before choosing a portable format.", "Prototype the product and evidence experience before scaling the brand shell.")
note("BELLI", "18:43-20:15", "Community acquisition", "Katie says she bought 20 Facebook groups totaling about 500K members.", "The useful lesson is access to dense conversation, not covert community control. Use transparent participation and compliant research.", "D", "Acquisition and membership totals unverified")
note("BELLI", "20:27-23:23", "Low-cost validation", "Home-kitchen production, a basic Shopify site, a retained day job, and Facebook communities powered early growth.", "Keep the MVP bounded and learn rapidly before adding fixed cost.")
note("BELLI", "27:32-31:18", "False PMF", "A narrow problem community made product-market fit look broader than it was; the first identity was too illness-centered and required a rebrand.", "Do not confuse dense niche enthusiasm with a scalable category or a durable name.")
note("BELLI", "32:08-33:31", "Retail pull", "Distinctive packaging and narrative attracted retail buyers, but early expansion was opportunistic.", "Channel interest is not proof of sell-through. Validate consumer, price, and shelf context.")
note("BELLI", "34:04-35:33", "Format risk", "A hot category did not guarantee the chosen format or aisle matched consumer intent.", "Test channel, format, price, and audience as separate hypotheses.")
note("BELLI", "36:02-39:57", "Founder fieldwork", "Katie personally stocked and inspected stores; Target exposed pricing, merchandising, and audience mismatch.", "Stay physically and analytically close to the buying surface.")
note("BELLI", "40:07-43:59", "Repositioning", "A year of R&D supported a move from sickness-framed bars toward aspirational fiber wellness.", "A narrow wedge can open the door, but the scalable category may need a more aspirational identity.")
note("BELLI", "44:04-47:33", "Channel-first launch", "Powders launched into 1,100 Walmart stores before DTC, initially with two SKUs.", "A focused assortment can be enough if the format and channel are validated.")
note("BELLI", "47:34-49:10", "Founder complement", "The founder describes relentless asking and doubling down; cofounders balance that intensity.", "Pair aggressive distribution instinct with operating and evidence discipline.")
note("BELLI", "49:11-50:35", "Ownable format", "A customer interaction video allegedly sold Walmart out, leading to permission for continued filming.", "Create a recurring format competitors cannot easily impersonate; obtain consent and verify impact.", "D", "Sellout and reach are company-reported")
note("BELLI", "50:36-51:29", "Creative volume", "Katie cleared 75 days, filmed thousands of interactions, and studied drop-off, likes, and reposts.", "Treat content as an instrumented production system, not occasional inspiration.")
note("BELLI", "51:30-53:47", "Research flywheel", "Content production created hundreds of direct customer and employee conversations.", "Make creator production double as structured qualitative research.")
note("BELLI", "54:39-54:54", "Context fit", "Founder admits the provocative early concept did not belong in every context.", "Use sharp language selectively and test reputational spillover before making it the master brand.")

# Mochi Health
note("MOCHI", "43:41-44:58", "Wedge to platform", "Mochi did not predict the GLP-1 wave; it began with an obesity-care gap and evolved toward a marketplace.", "Win a narrow problem before expanding into broad infrastructure.")
note("MOCHI", "45:02-45:20", "Network claim", "Founder claims more than 500 providers and around 100 pharmacies in the software ecosystem.", "Network breadth is useful only if choice, service, and data actually work; require definitions and active counts.", "D", "Current company article says 450+ providers")
note("MOCHI", "48:11-50:19", "Regulatory interpretation", "Founder explains compounding distinctions and marketing disputes.", "Treat as founder perspective, not legal guidance; use current primary sources for decisions.", "D", "Rapidly changing")
note("MOCHI", "52:41-54:03", "Trust proposition", "Real video visits, the same provider over time, formulary choice, and transparent cash pricing are the core differentiators.", "Continuity, choice, and transparent pricing are durable trust mechanics even outside clinical care.")
note("MOCHI", "54:54-56:28", "Competitor criticism", "Founder criticizes a competitor's provider flow and incentives.", "Do not repeat competitor assertions without independent verification.", "D", "Unverified opinion")
note("MOCHI", "57:17-57:32", "Founder tradeoff", "The founder left conventional medicine to build a system with broader reach.", "Mission can justify difficult operating choices, but measurable service outcomes must follow.")
note("MOCHI", "57:55-60:28", "Gray-market comparison", "The founder discusses rising peptide interest and broad safety differences between channels.", "Do not treat categorical founder opinions as proof of any specific seller's quality.", "D", "Opinion")
note("MOCHI", "60:28-61:58", "Testing bottleneck", "Pharmacy vetting and third-party testing capacity are framed as major bottlenecks.", "Secure test throughput, release SLAs, exception handling, and auditability before scaling inventory.")
note("MOCHI", "61:58-62:14", "Build versus buy", "At recording, Mochi had invested in a testing company rather than immediately building peptide production.", "Own or partner around the highest-control bottleneck before vertically integrating manufacturing.")
note("MOCHI", "63:18-63:33", "Future option", "Peptide expansion remained a future option at recording.", "Track what was planned versus what shipped; Mochi launched a longevity offer after the episode.")
note("MOCHI", "64:02-64:10", "Audience scope", "Founder says the target is ordinary Americans, not only elite biohackers.", "Translate technical trust for mainstream buyers without making unsupported outcome claims.")
note("MOCHI", "67:00-67:27", "Category fragmentation", "Peptides are described as a collection of mechanisms and use cases, not one coherent category.", "Avoid treating the catalog as one audience; build separate discovery and education paths.")
note("MOCHI", "67:27-68:03", "Capital discipline", "Minimal outside funding is positioned as protection from investor-driven audience bias.", "Choose capital that preserves the customer thesis and operating discipline.")


CLAIMS: list[dict] = []


def claim(
    claim_id: str,
    company: str,
    claim_text: str,
    claim_source: str,
    status: str,
    current_finding: str,
    evidence: str,
    confidence: int,
    impact: str,
    url: str,
) -> None:
    CLAIMS.append(
        {
            "Claim ID": claim_id,
            "Company": company,
            "Claim to evaluate": claim_text,
            "Where it came from": claim_source,
            "Verification status": status,
            "Current finding at 2026-07-26": current_finding,
            "Evidence class": EVIDENCE_LABELS[evidence],
            "Confidence (0-5)": confidence,
            "Decision impact": impact,
            "Verification link": url,
        }
    )


claim("CL-001", "Joi + Blokes", "More than $50M in revenue and essentially bootstrapped", "Founder interview and Entrepreneur", "Reported, not audited", "Entrepreneur repeated management's number; revenue period and definition remain unclear.", "B", 2, "Useful proof of possibility, not a forecast or valuation input.", "https://www.entrepreneur.com/building-a-business/they-built-a-50-million-business-before-glp-1s-became-mainstream-now-theyre-eyeing-this-new-health-frontier")
claim("CL-002", "Joi + Blokes", "Joi is the stronger growth engine and women bring male partners into Blokes", "Founder interview", "Founder-reported", "Architecture is directly visible; relative engine strength and referral behavior are not public metrics.", "D", 3, "Supports female-leading flagship plus shared backend, but measure it in Alex's cohorts.", "https://www.youtube.com/watch?v=aigF8IfN4Is&t=2046s")
claim("CL-003", "Joi + Blokes", "100K-patient footprint and fertility expansion", "Company announcements", "Company-reported; expansion verified", "HerMD acquisition and fertility leadership are public; patient count is not audited.", "C", 3, "Shows a path from focused fronts into adjacent life-stage services.", "https://www.businesswire.com/news/home/20260325782751/en/Joi-Blokes-Expands-Platform-with-Fertility-Division-Led-by-Serial-Founder-Gina-Bartasi")
claim("CL-004", "Peptira", "Business began 2025-06-02", "BBB business profile", "Corroborated", "BBB lists start date, LLC, and Benjamin Criss as contact/agent.", "B", 4, "Confirms this is a genuinely young operator.", "https://www.bbb.org/us/ne/chadron/profile/biological-products/peptira-llc-0714-1000080859")
claim("CL-005", "Peptira", "Top-15 US peptide company", "Founder interview", "Unsupported", "No independent ranking, revenue, traffic, or customer evidence was found.", "D", 1, "Ignore in competitor sizing.", "https://www.youtube.com/watch?v=Be58yiS6Vs4&t=1371s")
claim("CL-006", "Peptira", "Automatic replacement for lost shipments", "Founder interview", "Outdated / contradicted", "Live terms make delivered packages the customer risk, require evidence, and make remedies discretionary.", "C", 5, "Policies outrank interviews; never make service promises that terms and operations do not support.", "https://peptira.com/shipping-protection-terms/")
claim("CL-007", "Peptira", "RETA10-007 has a published third-party report", "Company COA library", "Historical document confirmed; custody not verified", "Peptira now marks RETA10-007 historical and RETA10-008 currently shipping. The historical Freedom Diagnostics report covers three submitted vials; sampling independence and inventory-wide representativeness remain unknown.", "C", 4, "Use batch reports, but label current versus historical and exactly what each report does and does not prove.", "https://drive.google.com/file/d/1i1tRAVLwnyypFHdKrg7SdBR7JsSk3Jj6/view?usp=drive_link")
claim("CL-008", "Peptide Price", "Community grew beyond the episode's 7,300 members", "Live Skool page", "Company-page observation", "Live community displayed roughly 11.8K members and $10/month at check.", "C", 4, "Comparison plus paid community is a durable discovery model.", "https://www.skool.com/peptide-price-9771/nooby-here")
claim("CL-009", "Transcend", "Ernest Colling remains the current owner/CEO", "Old podcast and lingering profiles", "Outdated", "Transcend introduced Lance Converse as new owner in March 2026; provider structure changed after December 2025.", "C", 5, "Do not use old founder interviews as present-company authority.", "https://www.linkedin.com/posts/transcendcompany_introducing-lance-converse-hes-stepping-activity-7443352442169090048-o66M")
claim("CL-010", "Transcend", "Creator scale broke support and pharmacy capacity", "Founder interview", "Plausible but unaudited", "Later complaints and provider disruption are consistent with service stress, but they do not prove the founder's exact causation or scale.", "D", 3, "Capacity-gate creator launches.", "https://www.youtube.com/watch?v=c6jUD2kyyDA&t=2226s")
claim("CL-011", "Transcend", "$1.5M foundation impact", "Founder interview", "Partly corroborated, amount unverified", "Public filings confirm the nonprofit and Colling role, not the dollar value.", "A", 2, "Mission metrics need auditable reporting.", "https://projects.propublica.org/nonprofits/organizations/920989536")
claim("CL-012", "Vita Bella", "$129/month membership plus $99 consultation", "Current official site", "Current company offer", "Canonical page confirms base pricing; California is priced separately.", "C", 5, "Useful recurring-access reference, not a direct RUO offer template.", "https://vitabella.com/membership/")
claim("CL-013", "Vita Bella", "10K+ patients and 15K+ consultations", "Current official homepage", "Company-reported", "Claims are live but unaudited.", "C", 3, "Do not use in sizing without independent evidence.", "https://vitabella.com/")
claim("CL-014", "Vita Bella", "State coverage, reviews, price, and delivery claims are fully consistent", "Current site pages", "Contradicted internally", "FAQ, search-indexed legacy pricing, terms, review figures, and delivery responsibility contain inconsistencies.", "C", 5, "Create one canonical facts source and audit every surface.", "https://vitabella.com/terms/")
claim("CL-015", "Anonymous growth case", "$500K-$600K monthly within five or six months", "Budai Media Ecom Show #320/#324/#325", "Anonymous agency claim", "No client identity, raw store data, processor settlements, or independent verification. Definitions change across episodes.", "D", 2, "Treat as an upside scenario only.", "https://shows.acast.com/theecomshow/episodes/325-budai-media-specials-if-i-had-to-scale-a-peptide-brand-t")
claim("CL-016", "Anonymous growth case", "Google more scalable than Meta for agency's peptide clients", "Ecom Show #324/#325", "Credible report of agency experience", "Repeated in freshest episode, but not a guarantee of policy approval or account durability.", "D", 4, "Test high-intent search while diversifying acquisition.", "https://shows.acast.com/theecomshow/episodes/325-budai-media-specials-if-i-had-to-scale-a-peptide-brand-t")
claim("CL-017", "Anonymous growth case", "70-80% gross margin and 1.3-1.4 break-even ROAS", "Agency episode", "Unsupported", "No cost definition; likely omits some testing, chargebacks, reserves, compliance, fulfillment, and account-loss costs.", "D", 1, "Build bottom-up unit economics instead.", "https://shows.acast.com/theecomshow/episodes/325-budai-media-specials-if-i-had-to-scale-a-peptide-brand-t")
claim("CL-018", "Remedora", "$200/month all-in-one telehealth launch platform", "Current official site", "Current company offer", "Site describes storefront, intake, provider, prescribing, pharmacy, payments, and support. Actual quality/coverage remains unverified.", "C", 3, "Useful partner benchmark; diligence every included function.", "https://remedora.com/")
claim("CL-019", "Remedora", "Retatrutide can be compounded in some states", "May 2026 podcast", "Outdated / contradicted", "FDA currently says retatrutide cannot be used in compounding under federal law.", "A", 5, "Remove from any current business-model assumption.", "https://www.fda.gov/drugs/drug-alerts-and-statements/fdas-concerns-unapproved-glp-1-drugs-used-weight-loss")
claim("CL-020", "Redline Biolabs", "Active founder-led store, multiple rails, and COA library", "Current site", "Confirmed footprint", "Store and documents are live; quality, customer count, and scale remain unverified.", "C", 4, "Good day-zero operating analog, not proof of a fast scaler.", "https://www.redlinebiolabs.com/lab-reports")
claim("CL-021", "Redline Biolabs", "COA on every batch", "Current site", "Internally inconsistent", "Some pages promise every-batch documentation while another says testing is pending.", "C", 4, "Never let marketing and release state disagree.", "https://www.redlinebiolabs.com/lab-reports")
claim("CL-022", "BioLongevity Labs", "650-member paid education community", "2024 founder interview", "Unverified historical claim", "Current media and research library are active, but member count was not independently verified.", "D", 1, "Education-led demand is useful; scale claim is not.", "https://mindpumppodcast.com/2500-weird-performance-enhancing-supplements-that-work-with-jay-campbell/")
claim("CL-023", "BioLongevity Labs", "80+ or 150+ products", "Current company pages", "Contradicted internally", "Site uses both figures while category inspection showed 53 peptide results.", "C", 2, "Catalog count is not traction and inconsistent proof harms trust.", "https://biolongevitylabs.com/product-category/peptides/")
claim("CL-024", "SheMed", "100K+ women served and $50M raise", "Company site and independent funding coverage", "Scale company-reported; funding reported", "Funding is reported; patient count remains private and unaudited.", "B", 4, "Strong evidence that female-first is not a small niche.", "https://www.mobihealthnews.com/news/shemed-raises-50m-expand-womens-glp-1-weight-management-platform")
claim("CL-025", "SheMed", "Clinical study validates the program", "Company site", "Results not yet published", "Study closed to recruitment and says results are expected in 2026; no published results found by cutoff.", "C", 4, "Never imply findings before publication.", "https://www.shemed.co.uk/clinical-study-why")
claim("CL-026", "SheMed / HeMed", "Female front can precede a male sibling on shared infrastructure", "Current company sites and company filing", "Visible current architecture", "HeMed launched in August 2025 using the same stated framework and infrastructure.", "C", 5, "Direct support for female-leading first, shared neutral operations, male front later.", "https://www.hemed.com/about")
claim("CL-027", "Arrae", "$100M+ cumulative revenue and 1.6M+ units", "Founder/company and Forbes profile", "Reported, not audited", "Retail expansion is observable; private financial totals remain management-reported.", "B", 3, "Creator system is transferable; financials are not planning inputs.", "https://www.forbes.com/sites/karineldor/2026/02/18/building-100m-wellness-brand-arrae-siff-haider-on-calculated-risk-and-relentless-standards/")
claim("CL-028", "Perelel", "$27M growth investment", "Company announcement and Fortune", "Corroborated", "Investment and approximate headcount were independently reported; revenue was not disclosed.", "B", 5, "High-confidence current analog for a female-leading identity brand.", "https://fortune.com/2025/11/20/perelel-prelude-growth-partners-27-million-beauty-wellness-supplements-women/")
claim("CL-029", "Perelel", "Profitable, doubled annual revenue, category-leading retention", "Company announcement", "Company-reported", "No raw financial or retention data are public.", "C", 3, "Use tactics, not claimed benchmarks.", "https://www.prnewswire.com/news-releases/perelel-secures-27m-growth-investment-from-prelude-growth-partners-to-reimagine-womens-health-302622022.html")
claim("CL-030", "BelliWelli", "405% revenue growth and 1B+ content views", "Company and US Chamber coverage", "Reported, not audited", "Current retail presence and $15.4M Series A are corroborated; growth/reach figures remain company-supplied.", "B", 3, "Content system is credible; scale metrics need caution.", "https://www.uschamber.com/co/good-company/launch-pad/health-startups-community-retail-growth")
claim("CL-031", "Mochi Health", "500K+ members, cash-flow positive, minimal funding", "Founder, company, and CEO Today", "Reported, not audited", "Multiple sources repeat the claims but originate with management.", "B", 3, "Useful operating philosophy, not valuation or market-size evidence.", "https://www.ceotodaymagazine.com/2026/07/myra-ahmad-cash-flow-positive-mochi-health/")
claim("CL-032", "Mochi Health", "500+ providers", "April 2026 interview", "Definition/timing discrepancy", "A July 2026 company article says 450+ providers. Could reflect active-count definitions or change.", "C", 3, "Define every network metric before reporting it.", "https://joinmochi.com/blog/how-mochi-is-building-a-provider-network-that-puts-patients-first")
claim("CL-033", "Mochi Health", "Not selling peptide programs at recording", "April 2026 interview", "True then, stale now", "Current longevity offering markets provider-supervised compounded NAD+ and sermorelin with disclosures.", "C", 5, "Rapidly changing category requires dated snapshots.", "https://joinmochi.com/longevity")
claim("CL-034", "Retatrutide", "Available as an approved public medication", "Market chatter and vendor marketing", "False as of cutoff", "Lilly says it remains investigational and not approved by any regulator; BLA planned Q1 2027.", "A", 5, "Do not describe it as approved or publicly available medicine.", "https://www.lilly.com/news/stories/what-to-know-about-retatrutide")
claim("CL-035", "FDA July peptide vote", "Six peptides became FDA-approved or immediately legal for compounding", "Post-vote social/media summaries", "False framing", "PCAC recommended six of seven for the 503A list. Recommendations are nonbinding and do not equal approval or immediate rule change.", "A", 5, "Keep regulatory event and commercial implication separate.", "https://www.fda.gov/advisory-committees/advisory-committee-calendar/july-23-24-2026-meeting-pharmacy-compounding-advisory-committee-07232026")
claim("CL-036", "RUO disclaimer", "'Research Use Only' controls intended use regardless of surrounding copy", "Common seller assumption", "Contradicted by current enforcement", "FDA's Gram Peptides warning says site evidence established human intended use despite RUO and not-for-human-consumption labels.", "A", 5, "Disclaimers cannot rescue contradictory content.", "https://www.fda.gov/inspections-compliance-enforcement-and-criminal-investigations/warning-letters/gram-peptides-721806-03312026")


CURRENT_CONTEXT = [
    {
        "Date": "2026-07-23 to 2026-07-24",
        "Topic": "FDA PCAC peptide votes",
        "What changed": "Advisory panel recommended BPC-157, KPV, TB-500, MOTS-c, Semax, and Epitalon for the 503A Bulks List; Emideltide/DSIP was rejected.",
        "What did NOT change": "The vote was nonbinding. It did not FDA-approve the substances and did not automatically make compounding or consumer sale lawful.",
        "Business implication": "Expect heightened attention and possibly more clinician/compounding competition. Do not treat headlines as an immediate green light.",
        "Evidence class": EVIDENCE_LABELS["A"],
        "Primary / current link": "https://www.fda.gov/advisory-committees/advisory-committee-calendar/july-23-24-2026-meeting-pharmacy-compounding-advisory-committee-07232026",
        "Next review": "When FDA posts final action or rulemaking",
    },
    {
        "Date": "2026-07-23",
        "Topic": "Retatrutide Phase 3",
        "What changed": "Lilly reported positive topline TRIUMPH-2 and TRIUMPH-3 results and plans a BLA submission in Q1 2027.",
        "What did NOT change": "Retatrutide remains investigational and is not FDA-approved or generally available from Lilly.",
        "Business implication": "Awareness and demand may accelerate, while manufacturer and regulator scrutiny can rise in parallel.",
        "Evidence class": EVIDENCE_LABELS["A"],
        "Primary / current link": "https://investor.lilly.com/news-releases/news-release-details/lillys-triple-agonist-retatrutide-successful-two-additional",
        "Next review": "Q1 2027 or earlier Lilly/FDA update",
    },
    {
        "Date": "2026-06-15 page update; checked 2026-07-26",
        "Topic": "Retatrutide and compounding",
        "What changed": "FDA now states explicitly that retatrutide and cagrilintide cannot be used in compounding under federal law.",
        "What did NOT change": "The July PCAC meeting concerned seven other substances, not retatrutide.",
        "Business implication": "Discard podcast-era assumptions that state-by-state interpretation creates a retatrutide compounding path.",
        "Evidence class": EVIDENCE_LABELS["A"],
        "Primary / current link": "https://www.fda.gov/drugs/drug-alerts-and-statements/fdas-concerns-unapproved-glp-1-drugs-used-weight-loss",
        "Next review": "Monthly and on any FDA update",
    },
    {
        "Date": "2026-03-31",
        "Topic": "RUO intended-use enforcement",
        "What changed": "FDA warning to Gram Peptides expressly rejected reliance on RUO/not-for-human-consumption labeling where website evidence established human intended use.",
        "What did NOT change": "A disclaimer remains one piece of evidence, not a safe harbor against contradictory copy and context.",
        "Business implication": "Creator scripts, FAQs, PDPs, support, emails, and community language must align with the actual lawful lane.",
        "Evidence class": EVIDENCE_LABELS["A"],
        "Primary / current link": "https://www.fda.gov/inspections-compliance-enforcement-and-criminal-investigations/warning-letters/gram-peptides-721806-03312026",
        "Next review": "On each new peptide warning letter",
    },
    {
        "Date": "2026-03-03",
        "Topic": "Telehealth GLP-1 marketing enforcement",
        "What changed": "FDA announced warning letters to 30 telehealth companies concerning allegedly illegal marketing of compounded GLP-1 products.",
        "What did NOT change": "A telehealth wrapper does not automatically make product marketing compliant.",
        "Business implication": "Do not assume migration from ecommerce to telehealth is a branding exercise; it changes licensed, clinical, pharmacy, and advertising obligations.",
        "Evidence class": EVIDENCE_LABELS["A"],
        "Primary / current link": "https://www.fda.gov/news-events/press-announcements/fda-warns-30-telehealth-companies-against-illegal-marketing-compounded-glp-1s",
        "Next review": "Quarterly enforcement review",
    },
    {
        "Date": "2026-07-21",
        "Topic": "Fast-growth agency's latest position",
        "What changed": "Budai Media updated its anonymous case to $600K in a month and said Google scaled more reliably than Meta.",
        "What did NOT change": "The client is still anonymous and the revenue, margin, and attribution claims remain unaudited.",
        "Business implication": "Treat high-intent search and lifecycle as tests, not guaranteed channels or benchmark economics.",
        "Evidence class": EVIDENCE_LABELS["D"],
        "Primary / current link": "https://shows.acast.com/theecomshow/episodes/325-budai-media-specials-if-i-had-to-scale-a-peptide-brand-t",
        "Next review": "Next case update or identified client evidence",
    },
    {
        "Date": "2026-03 to 2026-07",
        "Topic": "Transcend control and provider change",
        "What changed": "Transcend publicly introduced Lance Converse as new owner after a provider transition and active litigation.",
        "What did NOT change": "The 2024 Ernest Colling episode remains useful only as historical founder and scaling context.",
        "Business implication": "Verify current ownership, partners, complaint patterns, and service structure before learning from any old founder podcast.",
        "Evidence class": EVIDENCE_LABELS["C"],
        "Primary / current link": "https://transcendcompany.com/pages/change-of-provider-letter",
        "Next review": "Monthly until dispute and operating structure stabilize",
    },
    {
        "Date": "Checked 2026-07-26",
        "Topic": "Peptira service-policy drift",
        "What changed": "Live shipping and protection policies make remedies conditional and discretionary after delivery.",
        "What did NOT change": "The episode's service philosophy remains useful, but its automatic-replacement promise is not current.",
        "Business implication": "Maintain a dated promise register and automatically flag policy, support-script, and marketing conflicts.",
        "Evidence class": EVIDENCE_LABELS["C"],
        "Primary / current link": "https://peptira.com/shipping-protection-terms/",
        "Next review": "Monthly",
    },
    {
        "Date": "2025-12 to 2026-07",
        "Topic": "SheMed scale and sibling architecture",
        "What changed": "A four-month launch story developed into company-reported 100K+ women, reported $50M funding, and a male sibling brand on shared infrastructure.",
        "What did NOT change": "Private user metrics and study outcomes are not independently published.",
        "Business implication": "The female-leading front/shared-backend thesis has a current real-world precedent, but health claims are not transferable to RUO commerce.",
        "Evidence class": EVIDENCE_LABELS["B"],
        "Primary / current link": "https://www.mobihealthnews.com/news/shemed-raises-50m-expand-womens-glp-1-weight-management-platform",
        "Next review": "When study results publish",
    },
    {
        "Date": "2025-11 to 2026-07",
        "Topic": "Perelel growth investment",
        "What changed": "Perelel secured a confirmed $27M growth investment after five years and continued broadening across women's life stages.",
        "What did NOT change": "The company did not disclose revenue; profitability and retention remain company claims.",
        "Business implication": "Long creator arcs, expert editorial, stage-based taxonomy, and female-leading identity can scale without a female-exclusive catalog.",
        "Evidence class": EVIDENCE_LABELS["B"],
        "Primary / current link": "https://fortune.com/2025/11/20/perelel-prelude-growth-partners-27-million-beauty-wellness-supplements-women/",
        "Next review": "Next financing, retail, or annual update",
    },
    {
        "Date": "2026-04 to 2026-07",
        "Topic": "Mochi offering changed after interview",
        "What changed": "Mochi said it was not selling the discussed peptide programs at recording, then launched a longevity offering.",
        "What did NOT change": "Private member/provider claims remain unaudited; the offering is provider-supervised and not a template for RUO ecommerce.",
        "Business implication": "Use date-stamped snapshots because product and compliance claims can change within a quarter.",
        "Evidence class": EVIDENCE_LABELS["C"],
        "Primary / current link": "https://joinmochi.com/longevity",
        "Next review": "Monthly",
    },
    {
        "Date": "Checked 2026-07-26",
        "Topic": "Retail and creator analogs",
        "What changed": "Arrae, BelliWelli, and Perelel all expanded beyond their initial DTC wedge into larger assortments, retail, or life-stage systems.",
        "What did NOT change": "Founder-reported revenue and reach are not audited and do not prove the same tactics transfer to research compounds.",
        "Business implication": "Transfer seeding, cohort design, content systems, taxonomy, and customer research. Do not transfer consumer-use claims.",
        "Evidence class": EVIDENCE_LABELS["I"],
        "Primary / current link": "https://www.uschamber.com/co/good-company/launch-pad/health-startups-community-retail-growth",
        "Next review": "Semiannual",
    },
]


LESSONS = [
    {
        "Domain": "Brand architecture",
        "Lesson": "Launch a female-leading front on a neutral, reusable operating backbone.",
        "Evidence": "Joi + Blokes, SheMed/HeMed, Perelel",
        "Confidence": 5,
        "How to apply": "Let women dominate creators, stories, and default discovery. Keep technical PDPs and shared commerce accessible to male high-intent buyers; reserve a sibling male front if cohort economics justify it.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Conversion, CAC, AOV, second purchase, contribution, and referral rate by creator/buyer cohort",
    },
    {
        "Domain": "Positioning",
        "Lesson": "Differentiate the meaning and trust system, not merely the molecule or visual palette.",
        "Evidence": "Perelel, Arrae, Joi, Peptira",
        "Confidence": 5,
        "How to apply": "Build a rigorous identity, expert/research literacy, batch transparency, and creator narrative around a commodity catalog.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Direct traffic, branded search, creator-assisted conversion, price sensitivity, repeat rate",
    },
    {
        "Domain": "Assortment",
        "Lesson": "Use a presumed hero SKU as a hypothesis, not a permanent identity or proven revenue fact.",
        "Evidence": "Joi concentration story, Transcend, Vita Bella, internal Biologix popularity signal",
        "Confidence": 4,
        "How to apply": "Lead with the strongest observed demand signal while measuring product-level settled revenue, contribution, repeat interval, and cross-sell. Keep the company name broader.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Net revenue and contribution share by SKU; 30/60/90-day repeat by first product",
    },
    {
        "Domain": "Assortment",
        "Lesson": "Launch enough SKUs to prove the thesis, not enough to look impressive.",
        "Evidence": "Arrae, Perelel, BelliWelli, BioLongevity",
        "Confidence": 5,
        "How to apply": "Require every SKU to serve a defined research/customer job, basket logic, or acquisition demand. Add only after test, supply, content, and support readiness.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Sell-through, stock turns, support rate, margin, and repeat per SKU",
    },
    {
        "Domain": "Trust",
        "Lesson": "Batch-level proof is part of the product.",
        "Evidence": "Peptira, Redline, Mochi, Perelel",
        "Confidence": 5,
        "How to apply": "Expose lot ID, report date, methods, measured content, purity, relevant test scope, current/retired status, and report limitations near the buy decision.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "COA-open rate, conversion after COA view, support objections, lot-to-inventory match rate",
    },
    {
        "Domain": "Trust",
        "Lesson": "A published lab PDF is not the same as independent sampling and chain of custody.",
        "Evidence": "Peptira current report, Vita Bella unsupported competitor tests",
        "Confidence": 5,
        "How to apply": "Document who sampled, sealed, shipped, tested, released, and archived each batch. State what is not verified.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Percent of saleable inventory with complete custody record; release exceptions",
    },
    {
        "Domain": "Creator system",
        "Lesson": "Seed broadly and honestly, then pay creators who demonstrate authentic pull.",
        "Evidence": "Arrae, Perelel",
        "Confidence": 5,
        "How to apply": "Gift without posting obligation. Track organic mentions and audience quality; promote proven advocates into longer contracts.",
        "Priority": "P0 - Launch",
        "Metric / proof": "Organic post rate, qualified clicks, conversion, 60-day contribution, creator retention",
    },
    {
        "Domain": "Creator system",
        "Lesson": "Separate creator cohorts by job: awareness, trust, conversion, or community research.",
        "Evidence": "Perelel, Arrae, Redline",
        "Confidence": 5,
        "How to apply": "Give each cohort different briefs, compensation, landing pages, attribution windows, and success criteria.",
        "Priority": "P0 - Launch",
        "Metric / proof": "Incremental reach, assisted conversion, direct conversion, cohort contribution, branded search lift",
    },
    {
        "Domain": "Creator system",
        "Lesson": "Long arcs outperform rotating one-off endorsements in trust-heavy categories.",
        "Evidence": "Arrae, Perelel, Joi",
        "Confidence": 4,
        "How to apply": "Favor repeated, clearly disclosed relationships with continuity and evolving education.",
        "Priority": "P1 - Scale",
        "Metric / proof": "Repeat content performance, creator cohort retention, direct traffic, referral rate",
    },
    {
        "Domain": "Creator system",
        "Lesson": "Do not reward larger stacks or first-order spend.",
        "Evidence": "Vita Bella creator-incentive critique; compliance risk",
        "Confidence": 5,
        "How to apply": "Pay on qualified settled customers, compliant content, low refund/chargeback rate, and retained contribution.",
        "Priority": "P0 - Launch",
        "Metric / proof": "Net contribution at 60/90 days, refund and chargeback rate by creator",
    },
    {
        "Domain": "Content",
        "Lesson": "Make creator production double as structured customer research.",
        "Evidence": "BelliWelli, SheMed, Peptide Price",
        "Confidence": 5,
        "How to apply": "Code comments, objections, search terms, watch-drop points, support tickets, and post-purchase questions into a weekly insight log.",
        "Priority": "P0 - Launch",
        "Metric / proof": "Insights logged, experiments launched, conversion/support impact",
    },
    {
        "Domain": "Content",
        "Lesson": "Own a repeatable editorial format, not a feed of isolated product shots.",
        "Evidence": "BelliWelli, Arrae, Perelel",
        "Confidence": 5,
        "How to apply": "Create recurring proof, founder, lab, lot, creator, and category-literacy formats with a named owner.",
        "Priority": "P1 - Scale",
        "Metric / proof": "Completion rate, saves, shares, branded search, returning visitors",
    },
    {
        "Domain": "Acquisition",
        "Lesson": "Capture existing product-name intent before broad interruption advertising.",
        "Evidence": "Ecom Show #320/#324/#325",
        "Confidence": 3,
        "How to apply": "If eligible and approved, test high-intent search/Shopping with product and margin segmentation; reconcile platform attribution to settled orders.",
        "Priority": "P1 - Conditional",
        "Metric / proof": "Settled contribution per query/SKU, not platform ROAS alone",
    },
    {
        "Domain": "Acquisition",
        "Lesson": "Budget a bounded learning period with stopping rules.",
        "Evidence": "Ecom #320, Remedora",
        "Confidence": 4,
        "How to apply": "Predefine spend caps, signal thresholds, variables, decision windows, and kill/scale criteria.",
        "Priority": "P0 - Launch",
        "Metric / proof": "Experiment decision latency and percentage of tests with prewritten criteria",
    },
    {
        "Domain": "Lifecycle",
        "Lesson": "Triggered lifecycle belongs near launch, but attribution must be incremental.",
        "Evidence": "Ecom #320/#324/#325, Arrae, Vita Bella",
        "Confidence": 4,
        "How to apply": "Build welcome, abandonment, post-purchase, VIP, winback, and lawful replenishment flows; test holdouts where possible.",
        "Priority": "P0 - Launch",
        "Metric / proof": "Incremental conversion/repeat lift, unsubscribe, complaint, and deliverability",
    },
    {
        "Domain": "Data",
        "Lesson": "Own the event and identity graph because vendors can close accounts.",
        "Evidence": "Peptide Price TikTok bans, Ecom email shutdowns, payment instability",
        "Confidence": 5,
        "How to apply": "Track creator -> click -> order -> settlement -> fulfillment -> support -> repeat -> payout; export customer and event data continuously.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Event completeness, export freshness, restore-time test, unmatched-order rate",
    },
    {
        "Domain": "Data",
        "Lesson": "First-order ROAS is not the business.",
        "Evidence": "Remedora, Ecom case, Arrae operator split",
        "Confidence": 5,
        "How to apply": "Use settled contribution cohorts through 30/60/90 days, including refunds, chargebacks, reserves, testing, fulfillment, discounts, and creator payouts.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Contribution margin per visitor/customer by cohort and source",
    },
    {
        "Domain": "Operations",
        "Lesson": "Capacity-gate creators before demand arrives.",
        "Evidence": "Transcend, Peptira, Redline, Perelel",
        "Confidence": 5,
        "How to apply": "Require green status for saleable inventory, batch release, checkout, support staffing, fulfillment SLA, and exception ownership before each launch wave.",
        "Priority": "P0 - Launch",
        "Metric / proof": "Stockout, oversell, ship time, first response, resolution time, payment failure",
    },
    {
        "Domain": "Operations",
        "Lesson": "Founder service seeds trust but does not scale by itself.",
        "Evidence": "Peptira, Redline",
        "Confidence": 5,
        "How to apply": "Capture founder replies into macros, SOPs, escalation rules, and a staffed ticket system while founders remain visible on hard cases.",
        "Priority": "P0 - Launch",
        "Metric / proof": "Founder-handled share, first response, reopen rate, escalation rate",
    },
    {
        "Domain": "Operations",
        "Lesson": "Policies, marketing, support scripts, and checkout must share one canonical truth.",
        "Evidence": "Peptira replacement contradiction; Vita Bella state/price/review conflicts; Redline COA conflict",
        "Confidence": 5,
        "How to apply": "Create a dated promise register and automated/manual monthly surface audit.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Contradictions found, age of last audit, policy-linked support incidents",
    },
    {
        "Domain": "Payments",
        "Lesson": "Payment continuity is a launch dependency, not a back-office detail.",
        "Evidence": "Redline lost checkouts; Ecom case instability",
        "Confidence": 5,
        "How to apply": "Use honest underwriting, tested primary/backup rails, reserve reporting, decline telemetry, chargeback controls, and recovery playbooks.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Authorization rate, failure rate, reserve, chargebacks, recovery conversion",
    },
    {
        "Domain": "Supply",
        "Lesson": "Use staged supplier commitments and hard release gates.",
        "Evidence": "Redline supplier loss, Peptira test hold, Perelel MOQs",
        "Confidence": 5,
        "How to apply": "Verify counterparty, start with small POs, define specifications/acceptance, retain samples, and prohibit sale before complete release.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Supplier failure, batch rejection, release lead time, inventory at risk",
    },
    {
        "Domain": "Community",
        "Lesson": "Community is a research loop before it is a promotion channel.",
        "Evidence": "Peptide Price, BelliWelli, Arrae, SheMed",
        "Confidence": 5,
        "How to apply": "Listen transparently, classify objections and requests, and feed them into assortment, proof, content, and support decisions.",
        "Priority": "P1 - Scale",
        "Metric / proof": "Insight-to-experiment rate; member-to-customer and customer-to-referrer paths",
    },
    {
        "Domain": "Pricing",
        "Lesson": "Test same-product volume ladders before opaque bundles or silent subscriptions.",
        "Evidence": "Peptira, Ecom #324, Vita Bella",
        "Confidence": 4,
        "How to apply": "Show unit economics clearly, preserve choice, and require express recurring consent.",
        "Priority": "P1 - Test",
        "Metric / proof": "AOV, contribution, refund, repeat, and support by offer",
    },
    {
        "Domain": "Leadership",
        "Lesson": "Pair brand/culture authority with unit-economics and operations authority.",
        "Evidence": "Arrae, Perelel, Joi",
        "Confidence": 5,
        "How to apply": "Name accountable owners for brand/creator, growth/data, quality/supply, service/fulfillment, and financial control.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Decision latency, ownerless incidents, weekly scorecard completeness",
    },
    {
        "Domain": "Capital",
        "Lesson": "Raise to relieve proven constraints, not to purchase a narrative.",
        "Evidence": "Arrae, Perelel, Joi, Mochi",
        "Confidence": 4,
        "How to apply": "Tie capital to inventory turns, test capacity, people, working capital, and demonstrated acquisition/retention.",
        "Priority": "P2 - Later",
        "Metric / proof": "Use-of-funds milestones and runway sensitivity",
    },
    {
        "Domain": "Freshness",
        "Lesson": "Founder podcasts are dated snapshots, not current operating truth.",
        "Evidence": "Transcend ownership, Peptira policy, Mochi launch, Vita Bella California, Ecom updates",
        "Confidence": 5,
        "How to apply": "Every source row carries a current-check URL, freshness label, next-review date, and claim classification.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Percent of decision-critical sources refreshed within review window",
    },
    {
        "Domain": "Strategic risk",
        "Lesson": "Assume platform, processor, supplier, and regulatory concentration can break simultaneously.",
        "Evidence": "Ecom case, Peptide Price, Remedora prior ACA collapse, Transcend",
        "Confidence": 5,
        "How to apply": "Maintain a dependency register, backup plan, trigger, owner, and recovery-time test for every critical external system.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Single-point dependencies and tested recovery time",
    },
    {
        "Domain": "Operating lane",
        "Lesson": "Do not blend clinical, supplement, and RUO proof or claims.",
        "Evidence": "Joi, SheMed, Mochi, Vita Bella, BioLongevity",
        "Confidence": 5,
        "How to apply": "Separate sites/offers, role descriptions, evidence, support boundaries, and content rules for materially different models.",
        "Priority": "P0 - Foundational",
        "Metric / proof": "Content exceptions, support escalations, claim-review failures",
    },
    {
        "Domain": "Measurement",
        "Lesson": "Run the female-leading thesis as a controlled economic test, not a taste debate.",
        "Evidence": "SheMed visual testing, Joi architecture, Perelel, prior market research",
        "Confidence": 5,
        "How to apply": "Use matched female-leading and technical/general creator cohorts into differentiated landing journeys on shared commerce.",
        "Priority": "P0 - Launch",
        "Metric / proof": "Contribution per visitor, conversion, AOV, repeat, referral, and creative efficiency by cohort",
    },
]


ANTI_LESSONS = [
    {"Source / pattern": "Anonymous agency case", "Do not copy": "Treating $600K/month, margin, or ROAS claims as a forecast", "Why": "No client identity, settlements, raw cohort data, or independent verification", "Safer pattern": "Build bottom-up unit economics and use the case only to define hypotheses", "Risk level": "Critical"},
    {"Source / pattern": "Ecom Show #324/#325", "Do not copy": "Product renaming, alternate pages, or other platform-approval bypass tactics", "Why": "Evasion creates policy, account, legal, and trust exposure", "Safer pattern": "Seek explicit eligibility and approval; diversify lawful channels", "Risk level": "Critical"},
    {"Source / pattern": "RUO sellers", "Do not copy": "Human-use, reconstitution, mixing, dosing, protocol, or benefit content alongside RUO labels", "Why": "Current FDA enforcement looks at surrounding intended-use evidence", "Safer pattern": "Limit content to lawful research, testing, lot, documentation, shipping, and support topics", "Risk level": "Critical"},
    {"Source / pattern": "Remedora May 2026", "Do not copy": "Planning around retatrutide compounding", "Why": "Current FDA statement directly contradicts that assumption", "Safer pattern": "Use current primary-source review for every product and operating lane", "Risk level": "Critical"},
    {"Source / pattern": "Peptira", "Do not copy": "Automatic replacement, permanent pricing, or extreme response-time promises without durable operations", "Why": "Live policy now contradicts the early promise", "Safer pattern": "Publish only measured promises backed by the canonical policy", "Risk level": "High"},
    {"Source / pattern": "Peptira / Redline", "Do not copy": "Calling a vendor-submitted PDF complete proof of quality", "Why": "Sampling, chain of custody, lot coverage, and inventory linkage may be unknown", "Safer pattern": "Publish scope and limitations plus a complete custody record", "Risk level": "High"},
    {"Source / pattern": "Transcend", "Do not copy": "Scaling creators faster than support, fulfillment, or partner capacity", "Why": "Demand can break service and partner continuity", "Safer pattern": "Use launch-readiness gates and stop rules", "Risk level": "Critical"},
    {"Source / pattern": "Transcend", "Do not copy": "Relying on old founder titles or one practice/pharmacy relationship", "Why": "Ownership and provider structure materially changed", "Safer pattern": "Verify current control and contract continuity before acting", "Risk level": "High"},
    {"Source / pattern": "Vita Bella", "Do not copy": "Competitor-safety accusations without named reports and independent custody", "Why": "Claims are hard to defend and can mislead", "Safer pattern": "Compete on your own auditable proof", "Risk level": "High"},
    {"Source / pattern": "Vita Bella", "Do not copy": "Perfect-review, 24/7, wholesale-price, or rapid-delivery language that conflicts with terms", "Why": "Trust collapses when policies and marketing disagree", "Safer pattern": "Maintain a canonical promise register and surface audit", "Risk level": "High"},
    {"Source / pattern": "Ecom Show subscription advice", "Do not copy": "Preselected subscription or obscured recurring terms", "Why": "Consent and cancellation risk plus customer distrust", "Safer pattern": "Express choice, electronic confirmation, plain terms, and simple cancellation", "Risk level": "Critical"},
    {"Source / pattern": "Redline", "Do not copy": "Launching before payment, live inventory, testing state, refund rules, and service are stable", "Why": "It caused abandonment, overselling, and manual recovery", "Safer pattern": "Pass an end-to-end launch checklist under load", "Risk level": "Critical"},
    {"Source / pattern": "Redline", "Do not copy": "Mixing RUO restrictions with personal-use language in affiliate terms", "Why": "Internal contradiction weakens the operating position", "Safer pattern": "One reviewed creator policy with monitored scripts and enforcement", "Risk level": "Critical"},
    {"Source / pattern": "BioLongevity", "Do not copy": "Treating catalog count or announced formats as traction", "Why": "Site counts conflict and proposed pens were not found", "Safer pattern": "Report delivered products, current lot state, sell-through, and repeat behavior", "Risk level": "Medium"},
    {"Source / pattern": "BelliWelli", "Do not copy": "Covertly acquiring/influencing communities or filming without consent", "Why": "Platform, privacy, authenticity, and reputation risk", "Safer pattern": "Transparent participation and explicit consent", "Risk level": "High"},
    {"Source / pattern": "BelliWelli", "Do not copy": "Assuming niche-community excitement proves mass PMF", "Why": "Problem density can overstate addressable demand", "Safer pattern": "Validate channel, format, price, and audience separately", "Risk level": "High"},
    {"Source / pattern": "Arrae", "Do not copy": "One-off creators who rotate through competing products", "Why": "Audience trust and learning continuity stay weak", "Safer pattern": "Promote organic advocates into long-term, disclosed relationships", "Risk level": "Medium"},
    {"Source / pattern": "Perelel / SheMed / Mochi", "Do not copy": "Clinical, doctor-recommended, study, or patient-outcome language into RUO ecommerce", "Why": "The proof and licensed infrastructure are category-specific", "Safer pattern": "Transfer operating tactics, not clinical authority or claims", "Risk level": "Critical"},
    {"Source / pattern": "All founder podcasts", "Do not copy": "Repeating founder revenue, community, patient, or growth claims as facts", "Why": "Most private metrics are unaudited and often change definition", "Safer pattern": "Label claim provenance and seek independent or primary corroboration", "Risk level": "High"},
    {"Source / pattern": "All cases", "Do not copy": "Depending on one email vendor, processor, ad account, supplier, lab, creator, or regulatory interpretation", "Why": "Each dependency failed somewhere in the source set", "Safer pattern": "Dependency register, backup, trigger, owner, and tested recovery time", "Risk level": "Critical"},
]


ACTION_PLAN = [
    {"Window": "Now - Week 1", "Workstream": "Decision system", "Action": "Adopt this workbook's evidence labels and require each strategy claim to carry source, date, confidence, and current-check link.", "Owner": "Founder / strategy", "Dependency": "None", "Definition of done": "All decision-board claims have provenance and freshness fields", "Metric": "% critical claims with current verification", "Priority": "P0"},
    {"Window": "Now - Week 1", "Workstream": "Operating lane", "Action": "Write a one-page lane boundary for brand, product, content, support, creator scripts, and prohibited human-use guidance.", "Owner": "Legal/compliance + founder", "Dependency": "Qualified counsel", "Definition of done": "Approved rules are reflected in briefs, PDPs, FAQs, support, email, and community", "Metric": "Claim-review exceptions", "Priority": "P0"},
    {"Window": "Now - Week 1", "Workstream": "Brand architecture", "Action": "Lock neutral master name + female-leading flagship + technical PDP/shared-backend structure as the pilot hypothesis.", "Owner": "Founder + brand", "Dependency": "Current design pilot", "Definition of done": "One-page architecture names audiences, front doors, shared systems, and future male trigger", "Metric": "Decision signed; no contradictory site taxonomy", "Priority": "P0"},
    {"Window": "Now - Week 2", "Workstream": "Assortment", "Action": "Create a launch SKU gate: demand evidence, supplier diligence, complete test/custody, content, stock, margin, and support readiness.", "Owner": "Quality + merchandising", "Dependency": "Supplier and lab inputs", "Definition of done": "Every proposed SKU passes or is rejected with reason", "Metric": "% proposed SKUs passing all gates", "Priority": "P0"},
    {"Window": "Now - Week 2", "Workstream": "Trust", "Action": "Specify the lot record: lot ID, inventory mapping, sampling/custody, methods, measured content, purity, applicable scope, report date, status, and limitations.", "Owner": "Quality + product", "Dependency": "Lab workflow", "Definition of done": "Real current batch renders correctly on desktop/mobile and matches inventory", "Metric": "Lot-record completeness", "Priority": "P0"},
    {"Window": "Now - Week 2", "Workstream": "Promise control", "Action": "Create a canonical promise register for shipping, refunds, replacement, testing, origin, support hours, prices, reviews, and inventory.", "Owner": "Operations", "Dependency": "Policies and checkout", "Definition of done": "Marketing, policy, support, checkout, and creator briefs reconcile", "Metric": "Contradictions found per audit", "Priority": "P0"},
    {"Window": "Now - Week 2", "Workstream": "Payments", "Action": "Validate honest underwriting, primary and backup rails, reserves, declines, chargeback handling, and recovery messaging.", "Owner": "Finance + commerce", "Dependency": "Processor approval", "Definition of done": "End-to-end test orders pass; failure paths and owner are documented", "Metric": "Authorization/failure/chargeback rate", "Priority": "P0"},
    {"Window": "Now - Week 2", "Workstream": "Fulfillment", "Action": "Load-test saleable inventory, bundles/variants, stock holds, oversell prevention, label flow, tracking, and exception handling.", "Owner": "Operations", "Dependency": "Real inventory and carrier setup", "Definition of done": "Simulated launch wave completes with no unowned exception", "Metric": "Oversell, stockout, on-time ship, exception resolution", "Priority": "P0"},
    {"Window": "Week 2-3", "Workstream": "Data", "Action": "Instrument creator -> click -> order -> settlement -> fulfillment -> support -> repeat -> payout with stable IDs.", "Owner": "Growth ops / data", "Dependency": "Commerce and creator tooling", "Definition of done": "Test customer is traceable end to end and export restores successfully", "Metric": "Event completeness and unmatched order rate", "Priority": "P0"},
    {"Window": "Week 2-3", "Workstream": "Unit economics", "Action": "Build a settled-contribution model including COGS, testing, fulfillment, discounts, creator payout, processing, reserves, refunds, and chargebacks.", "Owner": "Finance", "Dependency": "Cost inputs", "Definition of done": "Every cohort and SKU has a contribution view at 0/30/60/90 days", "Metric": "Contribution per visitor/customer", "Priority": "P0"},
    {"Window": "Week 2-4", "Workstream": "Creator pilot", "Action": "Recruit a controlled 25-creator pilot split across female-leading, technical/general, awareness, and conversion roles.", "Owner": "Creator lead", "Dependency": "Approved scripts, landing journeys, inventory capacity", "Definition of done": "Each creator has role, source ID, code, brief, disclosure, compensation, and stop rule", "Metric": "Qualified traffic and contribution by creator", "Priority": "P0"},
    {"Window": "Week 2-4", "Workstream": "Seeding", "Action": "Run honest no-obligation seeding; identify organic advocates before offering long contracts.", "Owner": "Creator lead", "Dependency": "Seeding policy and fulfillment", "Definition of done": "Organic response is logged and top advocates are shortlisted by audience quality", "Metric": "Organic post and qualified-response rate", "Priority": "P1"},
    {"Window": "Week 3-4", "Workstream": "Landing test", "Action": "Create matched female-leading and technical/general acquisition journeys on the same backend and price structure.", "Owner": "Growth + product", "Dependency": "Design authority review for outward-facing UI", "Definition of done": "Both journeys are functional, approved, and differ only in the intended narrative variables", "Metric": "Contribution per visitor plus conversion/AOV/repeat", "Priority": "P0"},
    {"Window": "Week 3-4", "Workstream": "Lifecycle", "Action": "Implement welcome, abandonment, post-purchase documentation, service, VIP, winback, and lawful replenishment flows with consent controls.", "Owner": "Lifecycle", "Dependency": "Event data and policy", "Definition of done": "All flows tested; subscription is never preselected; exports work", "Metric": "Incremental lift, complaints, deliverability", "Priority": "P1"},
    {"Window": "Week 3-4", "Workstream": "Support", "Action": "Turn founder answers into macros, SOPs, escalation rules, and measured SLAs; keep founder visibility for hard cases.", "Owner": "Customer operations", "Dependency": "Promise register", "Definition of done": "Top objections and exceptions have owned playbooks", "Metric": "First response, resolution, reopen, founder share", "Priority": "P0"},
    {"Window": "Days 30-45", "Workstream": "Pilot launch", "Action": "Release creators in capacity-controlled waves, never all at once.", "Owner": "Growth + operations", "Dependency": "All launch gates green", "Definition of done": "Each wave is approved from live scorecard and can be paused in minutes", "Metric": "Service/fulfillment/payment thresholds stay green", "Priority": "P0"},
    {"Window": "Days 30-60", "Workstream": "Customer research", "Action": "Run weekly interviews and code creator comments, support tickets, search terms, checkout exits, and post-purchase questions.", "Owner": "Insights", "Dependency": "Consent and research template", "Definition of done": "Weekly insight memo produces ranked experiments", "Metric": "Insight-to-experiment and resolved-objection rate", "Priority": "P1"},
    {"Window": "Days 30-60", "Workstream": "Content engine", "Action": "Pilot recurring formats for founder visibility, lab/lot literacy, proof limitations, category research, and creator stories.", "Owner": "Content", "Dependency": "Reviewed content rules", "Definition of done": "Four repeatable formats have cadence, owner, and scorecard", "Metric": "Completion, saves, qualified clicks, assisted conversion", "Priority": "P1"},
    {"Window": "Days 30-60", "Workstream": "Offer testing", "Action": "Test one-unit versus simple same-product volume ladders; defer subscriptions until consent, service, and repeat behavior are proven.", "Owner": "Merchandising + finance", "Dependency": "Stable inventory and unit economics", "Definition of done": "Offer decision based on settled contribution and support burden", "Metric": "AOV, contribution, refund, repeat", "Priority": "P1"},
    {"Window": "Days 45-75", "Workstream": "Channel test", "Action": "Where explicitly eligible and approved, run a bounded high-intent search test segmented by product and margin.", "Owner": "Growth", "Dependency": "Platform approval and current policy review", "Definition of done": "Test has prewritten cap, attribution reconciliation, and kill/scale decision", "Metric": "Settled contribution per query/SKU", "Priority": "P2 conditional"},
    {"Window": "Days 60-75", "Workstream": "Cohort decision", "Action": "Compare female-leading versus technical/general cohorts on economics and behavior, not gross revenue.", "Owner": "Founder + finance + growth", "Dependency": "Sufficient settled orders", "Definition of done": "Written recommendation to intensify female lead, preserve dual funnels, or adjust", "Metric": "Contribution, repeat, referral, CAC, AOV by cohort", "Priority": "P0"},
    {"Window": "Days 60-90", "Workstream": "Creator contracts", "Action": "Promote proven organic or pilot creators into longer relationships with KPI-specific compensation.", "Owner": "Creator lead + finance", "Dependency": "Cohort contribution", "Definition of done": "No long contract is based solely on reach or one post", "Metric": "90-day creator contribution and audience quality", "Priority": "P1"},
    {"Window": "Days 60-90", "Workstream": "Dependency resilience", "Action": "Run recovery drills for processor, email, supplier, lab, shipping, and creator-channel failure.", "Owner": "Operations + engineering", "Dependency": "Backup paths", "Definition of done": "Each dependency has tested RTO, owner, trigger, and customer communication", "Metric": "Recovery time and data loss", "Priority": "P0"},
    {"Window": "Day 90", "Workstream": "Go / no-go", "Action": "Decide whether to scale, narrow, or stop using settled contribution, repeat behavior, trust/support signals, and dependency risk.", "Owner": "Founder / board", "Dependency": "Complete cohort and incident data", "Definition of done": "Decision memo includes evidence, confidence, downside case, and next capital commitment", "Metric": "Decision quality, not activity volume", "Priority": "P0"},
]


WATCHLIST = [
    {"Subject": "FDA PCAC six-peptide recommendations", "What to watch": "Final FDA action, proposed rule, guidance, or enforcement discretion", "Current state": "Nonbinding panel recommendation only", "Trigger": "Any Federal Register/FDA final action", "Cadence": "Weekly", "Owner": "Compliance", "Source": "https://www.fda.gov/advisory-committees/advisory-committee-calendar/july-23-24-2026-meeting-pharmacy-compounding-advisory-committee-07232026"},
    {"Subject": "Retatrutide", "What to watch": "Lilly BLA timing, trial publications, warning letters, litigation, and FDA statements", "Current state": "Investigational; BLA planned Q1 2027; FDA says cannot be compounded", "Trigger": "Regulatory filing or status change", "Cadence": "Weekly", "Owner": "Compliance / strategy", "Source": "https://www.lilly.com/news/stories/what-to-know-about-retatrutide"},
    {"Subject": "Gram Peptides and similar enforcement", "What to watch": "How FDA characterizes intended-use evidence around RUO sites", "Current state": "Current warning letter says disclaimers did not override surrounding human-use evidence", "Trigger": "New peptide warning letter", "Cadence": "Monthly", "Owner": "Compliance", "Source": "https://www.fda.gov/inspections-compliance-enforcement-and-criminal-investigations/warning-letters/gram-peptides-721806-03312026"},
    {"Subject": "Anonymous Budai client", "What to watch": "Client identity, processor evidence, settled revenue, refunds, repeat, and future case updates", "Current state": "$600K monthly remains anonymous agency claim", "Trigger": "Named client or raw evidence", "Cadence": "Monthly", "Owner": "Growth", "Source": "https://shows.acast.com/theecomshow/episodes/325-budai-media-specials-if-i-had-to-scale-a-peptide-brand-t"},
    {"Subject": "Joi + Blokes", "What to watch": "Fertility launch, HerMD integration, current leadership, and audience/revenue mix", "Current state": "Active and expanding; private scale unaudited", "Trigger": "New funding, acquisition, or founder interview", "Cadence": "Quarterly", "Owner": "Strategy", "Source": "https://joiandblokes.com/"},
    {"Subject": "Peptira / Peptide Price", "What to watch": "Catalog, batch-report currency, community size, affiliate disclosure, policy changes, and service claims", "Current state": "Active; shipping policy drift found", "Trigger": "New podcast or material policy/catalog change", "Cadence": "Monthly", "Owner": "Competitive intel", "Source": "https://peptira.com/test-results/"},
    {"Subject": "Transcend", "What to watch": "Ownership, provider continuity, litigation, complaint trend, and service recovery", "Current state": "New owner and changed provider structure", "Trigger": "Court disposition or formal operating update", "Cadence": "Monthly", "Owner": "Strategy", "Source": "https://transcendcompany.com/pages/change-of-provider-letter"},
    {"Subject": "SheMed", "What to watch": "Publication of clinical-study results, US expansion, HeMed shared-infrastructure economics", "Current state": "Study results not published by cutoff", "Trigger": "Peer-reviewed/public result or expansion update", "Cadence": "Monthly", "Owner": "Brand / strategy", "Source": "https://www.shemed.co.uk/clinical-study-why"},
    {"Subject": "Perelel", "What to watch": "Creator-cohort changes, retail expansion, retention disclosures, and male-product treatment", "Current state": "$27M growth investment confirmed", "Trigger": "New founder interview or annual update", "Cadence": "Quarterly", "Owner": "Brand / creator", "Source": "https://perelelhealth.com/"},
    {"Subject": "Arrae", "What to watch": "Creator mix, subscription mechanics, retail expansion, and verified scale", "Current state": "Company-reported $100M cumulative revenue", "Trigger": "Independent financial or channel update", "Cadence": "Quarterly", "Owner": "Creator / lifecycle", "Source": "https://www.arrae.com/pages/about-us"},
    {"Subject": "BelliWelli", "What to watch": "Repeatable content format, retailer sell-through, subscription, and community transparency", "Current state": "Retail footprint and funding confirmed; growth/reach private", "Trigger": "New channel or financial disclosure", "Cadence": "Quarterly", "Owner": "Content / community", "Source": "https://belliwelli.com/"},
    {"Subject": "Mochi Health", "What to watch": "Longevity offering, testing ownership, provider count, litigation, and price/choice model", "Current state": "Offering changed within one quarter; private scale claims vary", "Trigger": "Product, court, or company metric update", "Cadence": "Monthly", "Owner": "Strategy", "Source": "https://joinmochi.com/longevity"},
]


EXECUTIVE_DECISIONS = [
    {
        "Question": "Female-leading or unisex?",
        "Answer": "Female-leading flagship, neutral name, technical/shared commerce, not female-exclusive.",
        "Evidence": "Joi + Blokes, SheMed/HeMed, Perelel",
        "Confidence": "High",
        "Why": "These operators show that distinct audience meaning can sit on shared infrastructure. Perelel even carries a men's product inside a female-leading master brand.",
        "First proof": "Matched female-leading vs technical/general creator cohort test",
    },
    {
        "Question": "What is the moat?",
        "Answer": "Creator relationships + batch proof + owned event/identity graph + dependable service.",
        "Evidence": "Arrae, Perelel, Peptira, Redline, Ecom cases",
        "Confidence": "High",
        "Why": "Molecules, catalogs, and storefront software commoditize. Trust, continuity, attribution, and response quality compound.",
        "First proof": "Track creator -> settlement -> fulfillment -> support -> repeat -> payout",
    },
    {
        "Question": "Should one product lead?",
        "Answer": "Use the strongest observed demand signal as a launch hypothesis, but keep the brand and economics broader.",
        "Evidence": "Joi and Vita Bella concentration stories; internal popularity signal",
        "Confidence": "Medium-high",
        "Why": "A hero offer can concentrate acquisition while hiding operational and strategic concentration.",
        "First proof": "Product-level settled revenue, contribution, repeat interval, and cross-sell",
    },
    {
        "Question": "How should creators be recruited?",
        "Answer": "Controlled seeding, then long-term deals with proven organic advocates; separate awareness and conversion cohorts.",
        "Evidence": "Arrae and Perelel",
        "Confidence": "High",
        "Why": "Repeated genuine affinity is more credible and more learnable than rotating one-off endorsements.",
        "First proof": "25-creator capacity-controlled pilot with role-specific KPIs",
    },
    {
        "Question": "How much should growth stories influence the plan?",
        "Answer": "Very little until independently reconciled.",
        "Evidence": "Anonymous Ecom case, shifting definitions, unaudited founder claims",
        "Confidence": "High",
        "Why": "The most exciting numbers have the weakest evidence. They define upside hypotheses, not budgets.",
        "First proof": "Bottom-up contribution model and raw settled cohort evidence",
    },
    {
        "Question": "What fails first?",
        "Answer": "Payments, support, inventory, testing throughput, fulfillment, or a partner handoff before demand itself.",
        "Evidence": "Transcend, Redline, Peptira, Ecom case",
        "Confidence": "High",
        "Why": "Every exact-category operator described operational or vendor failure during growth.",
        "First proof": "Launch-readiness gates, stop rules, and recovery drills",
    },
    {
        "Question": "How current is the opportunity?",
        "Answer": "Demand visibility is rising, but the operating environment is changing within weeks.",
        "Evidence": "July PCAC vote, July Lilly data, EC325, Mochi change, Transcend change",
        "Confidence": "High",
        "Why": "Current information increases attention while also invalidating earlier assumptions. A dated monitoring system is mandatory.",
        "First proof": "Weekly current-context review and claim-expiry register",
    },
]


FAST_QUEUE = [
    ("1", "Ecom Show #325", "Full 15:15", "Freshest exact scaling update; focus on channel reality, lifecycle, payment instability, and the 'near-term window' warning.", "https://shows.acast.com/theecomshow/episodes/325-budai-media-specials-if-i-had-to-scale-a-peptide-brand-t"),
    ("2", "Joi + Blokes", "34:06-37:51; 48:53-54:36", "Female/male fronts on one backend, women as stronger engine, scale/capital, and adjacency.", "https://www.youtube.com/watch?v=aigF8IfN4Is&t=2046s"),
    ("3", "Peptira + Peptide Price", "14:38-31:38", "Six-week launch sequence, batch labels, COAs, volume ladder, support promises, and test-before-sale gate.", "https://www.youtube.com/watch?v=Be58yiS6Vs4&t=878s"),
    ("4", "Perelel", "11:27-20:30", "Identity brand, unique capability, long creator arcs, seeding, message order, and creator cohorts.", "https://podcasts.apple.com/us/podcast/success-almost-broke-her-this-recovering-girlboss-built/id1593458926?i=1000743844412"),
    ("5", "SheMed", "09:34-28:48", "Role boundaries, shareable education, ongoing support, channel mix, muted-color conversion learning, and real buyer age.", "https://www.buzzsprout.com/850540/episodes/17034792-the-weightloss-wonders"),
    ("6", "Redline Biolabs", "17:30-20:45; 35:45-47:15", "Founders + COAs, payment abandonment, supplier loss, oversell recovery, attribution codes, and early service.", "https://www.youtube.com/watch?v=JQSJYACbOCU&t=1050s"),
    ("7", "Arrae", "15:45-29:35; 34:28-35:50", "Soft launch, no-obligation seeding, promote organic love, customer-as-creator, and founder role split.", "https://podcasts.apple.com/us/podcast/siffat-haider-co-founder-co-ceo-of-arrae-what-makes/id1432195522?i=1000611429597"),
    ("8", "Transcend", "35:07-40:14", "Historical scale-break caution: creators, staffing, calls, patient capacity, and pharmacy inventory.", "https://www.youtube.com/watch?v=c6jUD2kyyDA&t=2107s"),
]


def build_web_dataset() -> Path:
    """Emit the browser-ready dataset from the same records as the workbook."""
    payload = {
        "asOf": AS_OF,
        "stats": {
            "sources": len(SOURCES),
            "notes": len(NOTES),
            "claims": len(CLAIMS),
            "currentContext": len(CURRENT_CONTEXT),
            "lessons": len(LESSONS),
            "antiLessons": len(ANTI_LESSONS),
            "actions": len(ACTION_PLAN),
            "watchlist": len(WATCHLIST),
        },
        "sources": SOURCES,
        "notes": NOTES,
        "claims": CLAIMS,
        "currentContext": CURRENT_CONTEXT,
        "lessons": LESSONS,
        "antiLessons": ANTI_LESSONS,
        "actions": ACTION_PLAN,
        "watchlist": WATCHLIST,
        "executiveDecisions": EXECUTIVE_DECISIONS,
        "fastQueue": FAST_QUEUE,
    }
    encoded = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    WEB_OUTPUT.write_text(
        "window.PEPTIDE_MEDIA_INTELLIGENCE = " + encoded + ";\n",
        encoding="utf-8",
    )
    return WEB_OUTPUT


def valid_url(value: str) -> bool:
    if not value:
        return True
    parsed = urlparse(value)
    return parsed.scheme in {"http", "https"} and bool(parsed.netloc)


def add_hyperlink(cell, url: str, label: str | None = None) -> None:
    if url and valid_url(url):
        if label:
            cell.value = label
        cell.hyperlink = url
        cell.font = Font(name="Aptos", size=10, color=COBALT, underline="single")


def apply_value_fill(cell) -> None:
    value = str(cell.value or "")
    lower = value.lower()
    if value.startswith("A -"):
        cell.fill = PatternFill("solid", fgColor=SAGE_LIGHT)
        cell.font = Font(name="Aptos", size=10, color=SAGE, bold=True)
    elif value.startswith("B -"):
        cell.fill = PatternFill("solid", fgColor=COBALT_LIGHT)
        cell.font = Font(name="Aptos", size=10, color=COBALT, bold=True)
    elif value.startswith("C -"):
        cell.fill = PatternFill("solid", fgColor=GOLD_LIGHT)
        cell.font = Font(name="Aptos", size=10, color=ESPRESSO, bold=True)
    elif value.startswith("D -"):
        cell.fill = PatternFill("solid", fgColor=CLAY_LIGHT)
        cell.font = Font(name="Aptos", size=10, color=CLAY, bold=True)
    elif value.startswith("E -") or "critical" in lower or "false" in lower or "contradicted" in lower:
        cell.fill = PatternFill("solid", fgColor=RED_LIGHT)
        cell.font = Font(name="Aptos", size=10, color=RED, bold=True)
    elif "p0" in lower or "start here" in lower or "high" == lower:
        cell.fill = PatternFill("solid", fgColor=SAGE_LIGHT)
        cell.font = Font(name="Aptos", size=10, color=SAGE, bold=True)
    elif "materially changed" in lower or "outdated" in lower or "unsupported" in lower:
        cell.fill = PatternFill("solid", fgColor=RED_LIGHT)
        cell.font = Font(name="Aptos", size=10, color=RED, bold=True)
    elif "very current" in lower or "current; high" in lower or "corroborated" in lower or "confirmed" in lower:
        cell.fill = PatternFill("solid", fgColor=COBALT_LIGHT)


def add_tabular_sheet(
    wb: Workbook,
    title: str,
    rows: list[dict],
    subtitle: str,
    widths: dict[str, float] | None = None,
    tab_color: str = SAGE,
) -> None:
    ws = wb.create_sheet(title)
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = tab_color
    if not rows:
        ws["A1"] = title
        return
    headers = list(rows[0].keys())
    last_col = get_column_letter(len(headers))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = title.upper()
    ws["A1"].fill = PatternFill("solid", fgColor=INK)
    ws["A1"].font = Font(name="Aptos Display", size=19, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 34
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = subtitle
    ws["A2"].fill = PatternFill("solid", fgColor=CREAM)
    ws["A2"].font = Font(name="Aptos", size=10, color=ESPRESSO, italic=True)
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 34
    ws.merge_cells(f"A3:{last_col}3")
    ws["A3"] = f"Current check: {AS_OF}  |  Evidence is classified, not assumed. This is business research, not medical or legal advice."
    ws["A3"].font = Font(name="Aptos", size=9, color=MUTED)
    ws["A3"].alignment = Alignment(vertical="center")
    ws.row_dimensions[3].height = 22
    header_row = 5
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(header_row, col_idx, header)
        cell.fill = PatternFill("solid", fgColor=ESPRESSO)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
        cell.border = BORDER
    ws.row_dimensions[header_row].height = 42
    for row_idx, item in enumerate(rows, header_row + 1):
        max_len = 0
        for col_idx, header in enumerate(headers, 1):
            value = item.get(header, "")
            cell = ws.cell(row_idx, col_idx, value)
            cell.font = Font(name="Aptos", size=10, color=INK)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=PORCELAIN)
            apply_value_fill(cell)
            if "link" in header.lower() or header in {"Source", "Primary / current link", "Verification link", "Media link", "Current-check link", "Timestamp link"}:
                add_hyperlink(cell, str(value), "Open source")
            max_len = max(max_len, len(str(value)))
        ws.row_dimensions[row_idx].height = min(108, max(30, 15 + 0.30 * max_len))
    ref = f"A{header_row}:{last_col}{header_row + len(rows)}"
    table_name = "T_" + "".join(ch for ch in title if ch.isalnum())[:24]
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.freeze_panes = f"A{header_row + 1}"
    ws.auto_filter.ref = ref
    default_width = 18
    for col_idx, header in enumerate(headers, 1):
        letter = get_column_letter(col_idx)
        chosen = (widths or {}).get(header)
        if chosen is None:
            if "link" in header.lower() or header == "Source":
                chosen = 14
            elif header in {"ID", "Source ID", "Claim ID", "Published", "Duration", "Date", "Cadence", "Owner"}:
                chosen = 14
            elif len(header) > 22:
                chosen = 28
            else:
                chosen = default_width
        ws.column_dimensions[letter].width = chosen
    ws.auto_filter.ref = ref
    ws.sheet_view.zoomScale = 85


def build_start_here(wb: Workbook) -> None:
    ws = wb.active
    ws.title = "START HERE"
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = GOLD
    widths = {
        "A": 7,
        "B": 24,
        "C": 24,
        "D": 24,
        "E": 24,
        "F": 18,
        "G": 18,
        "H": 18,
        "I": 3,
        "J": 25,
        "K": 22,
        "L": 20,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    ws.merge_cells("A1:L2")
    ws["A1"] = "PEPTIDE FOUNDER MEDIA INTELLIGENCE"
    ws["A1"].fill = PatternFill("solid", fgColor=INK)
    ws["A1"].font = Font(name="Aptos Display", size=24, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.merge_cells("A3:L3")
    ws["A3"] = "Recent founder and operator media, fully separated from present-day verification and translated into a Biologix operating playbook."
    ws["A3"].fill = PatternFill("solid", fgColor=CREAM)
    ws["A3"].font = Font(name="Aptos", size=11, color=ESPRESSO, italic=True)
    ws["A3"].alignment = Alignment(vertical="center")
    ws.row_dimensions[3].height = 30

    metadata = [
        ("As of", AS_OF),
        ("Media sources", len(SOURCES)),
        ("Timestamped extractions", len(NOTES)),
        ("Material claims checked", len(CLAIMS)),
        ("Current-context updates", len(CURRENT_CONTEXT)),
        ("Action rows", len(ACTION_PLAN)),
    ]
    for row, (label, value) in enumerate(metadata, 5):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.cell(row, 1, label).font = Font(name="Aptos", bold=True, color=MUTED)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=4)
        ws.cell(row, 3, value).font = Font(name="Aptos", bold=True, color=INK)
    ws.merge_cells("A12:H12")
    ws["A12"] = "THE DECISION READOUT"
    ws["A12"].fill = PatternFill("solid", fgColor=ESPRESSO)
    ws["A12"].font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    ws["A12"].alignment = Alignment(vertical="center")
    decisions = [
        ("01", "Architecture", "Female-leading flagship, neutral name, technical/shared commerce, and a future male sibling only if cohort economics earn it."),
        ("02", "Commercial hypothesis", "Use the strongest observed demand signal as a launch wedge, but do not name the company after one product or assume popularity equals settled revenue."),
        ("03", "Moat", "Creators + batch proof + owned event graph + dependable service. The storefront software and catalog are not the moat."),
        ("04", "Launch method", "Controlled 25-creator pilot, split by female-leading vs technical/general and awareness vs conversion roles; release in capacity-gated waves."),
        ("05", "Evidence discipline", "Do not forecast from anonymous $600K/month claims. Require settled contribution and 30/60/90-day cohorts."),
        ("06", "Freshness warning", "Current facts already invalidate old media: Transcend changed control, Peptira changed remedies, Mochi changed offering, and retatrutide compounding claims are obsolete."),
    ]
    for idx, (num, label, text) in enumerate(decisions, 13):
        ws.cell(idx, 1, num).font = Font(name="Aptos Display", size=18, bold=True, color=GOLD)
        ws.cell(idx, 2, label).font = Font(name="Aptos", size=10, bold=True, color=ESPRESSO)
        ws.merge_cells(start_row=idx, start_column=3, end_row=idx, end_column=8)
        ws.cell(idx, 3, text).font = Font(name="Aptos", size=10, color=INK)
        ws.cell(idx, 3).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[idx].height = 48
        for col in range(1, 9):
            ws.cell(idx, col).border = Border(bottom=THIN_STONE)

    ws.merge_cells("A20:H20")
    ws["A20"] = "FASTEST LISTENING QUEUE"
    ws["A20"].fill = PatternFill("solid", fgColor=ESPRESSO)
    ws["A20"].font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    queue_headers = ["Rank", "Source", "Listen", "Why", "Link"]
    for col, header in enumerate(queue_headers, 1):
        cell = ws.cell(21, col, header)
        cell.fill = PatternFill("solid", fgColor=STONE)
        cell.font = Font(name="Aptos", bold=True, color=INK)
        cell.border = BORDER
        cell.alignment = Alignment(wrap_text=True)
    ws.merge_cells("D21:G21")
    ws.merge_cells("H21:H21")
    # Re-create the visual header after merging.
    ws["D21"] = "Why"
    ws["H21"] = "Link"
    for row_idx, (rank, name, listen, why, url) in enumerate(FAST_QUEUE, 22):
        ws.cell(row_idx, 1, rank)
        ws.cell(row_idx, 2, name)
        ws.cell(row_idx, 3, listen)
        ws.merge_cells(start_row=row_idx, start_column=4, end_row=row_idx, end_column=7)
        ws.cell(row_idx, 4, why)
        add_hyperlink(ws.cell(row_idx, 8), url, "Open")
        for col in range(1, 9):
            cell = ws.cell(row_idx, col)
            cell.border = BORDER
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=PORCELAIN)
        ws.row_dimensions[row_idx].height = 52

    ws.merge_cells("J5:L5")
    ws["J5"] = "EVIDENCE KEY"
    ws["J5"].fill = PatternFill("solid", fgColor=ESPRESSO)
    ws["J5"].font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    for idx, code in enumerate(["A", "B", "C", "D", "E", "I"], 6):
        ws.cell(idx, 10, code).font = Font(name="Aptos", bold=True, color=INK)
        ws.merge_cells(start_row=idx, start_column=11, end_row=idx, end_column=12)
        ws.cell(idx, 11, EVIDENCE_LABELS[code].split(" - ", 1)[1])
        ws.cell(idx, 11).alignment = Alignment(wrap_text=True)
        apply_value_fill(ws.cell(idx, 10))
        ws.row_dimensions[idx].height = 32

    ws.merge_cells("J13:L13")
    ws["J13"] = "CURRENT ALERTS"
    ws["J13"].fill = PatternFill("solid", fgColor=RED)
    ws["J13"].font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    alerts = [
        ("Retatrutide", "Investigational, not FDA-approved; FDA says it cannot be compounded."),
        ("July PCAC vote", "Six recommendations are nonbinding, not approvals or immediate legal change."),
        ("RUO labels", "Current FDA warning shows disclaimers do not override contradictory intended-use evidence."),
        ("Transcend", "2024 interview is historical; new owner/provider structure in 2026."),
        ("Peptira", "Live shipping remedy is stricter than early founder promise."),
    ]
    for row, (label, text) in enumerate(alerts, 14):
        ws.cell(row, 10, label).font = Font(name="Aptos", bold=True, color=RED)
        ws.merge_cells(start_row=row, start_column=11, end_row=row, end_column=12)
        ws.cell(row, 11, text)
        ws.cell(row, 11).alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[row].height = 48

    lane_counts = Counter(item["Lane"] for item in SOURCES)
    ws["J21"] = "Source lane"
    ws["K21"] = "Count"
    for cell in ws[21][9:11]:
        cell.fill = PatternFill("solid", fgColor=STONE)
        cell.font = Font(name="Aptos", bold=True, color=INK)
    for row, (lane, count) in enumerate(lane_counts.items(), 22):
        ws.cell(row, 10, lane)
        ws.cell(row, 11, count)
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Source mix"
    chart.height = 6.2
    chart.width = 11
    data = Reference(ws, min_col=11, min_row=21, max_row=21 + len(lane_counts))
    cats = Reference(ws, min_col=10, min_row=22, max_row=21 + len(lane_counts))
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    chart.x_axis.title = "Sources"
    ws.add_chart(chart, "J29")

    ws.merge_cells("A32:H34")
    ws["A32"] = (
        "Use this workbook as a dated operator-intelligence system. It intentionally excludes dosing and personal-use instructions. "
        "Founder and agency claims are preserved because they reveal operating hypotheses, but they are never upgraded to facts without verification."
    )
    ws["A32"].fill = PatternFill("solid", fgColor=GOLD_LIGHT)
    ws["A32"].font = Font(name="Aptos", size=10, color=ESPRESSO, italic=True)
    ws["A32"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A5"
    ws.sheet_view.zoomScale = 90


def build_executive_sheet(wb: Workbook) -> None:
    ws = wb.create_sheet("EXECUTIVE SYNTHESIS")
    ws.sheet_view.showGridLines = False
    ws.sheet_properties.tabColor = SAGE
    widths = [23, 38, 28, 15, 48, 34]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.merge_cells("A1:F1")
    ws["A1"] = "EXECUTIVE SYNTHESIS"
    ws["A1"].fill = PatternFill("solid", fgColor=INK)
    ws["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 36
    ws.merge_cells("A2:F2")
    ws["A2"] = "The transferable lessons, stripped of hype and updated through July 26, 2026."
    ws["A2"].fill = PatternFill("solid", fgColor=CREAM)
    ws["A2"].font = Font(name="Aptos", size=10, italic=True, color=ESPRESSO)
    headers = list(EXECUTIVE_DECISIONS[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(4, col, header)
        cell.fill = PatternFill("solid", fgColor=ESPRESSO)
        cell.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        cell.alignment = Alignment(wrap_text=True)
        cell.border = BORDER
    for row, decision in enumerate(EXECUTIVE_DECISIONS, 5):
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row, col, decision[header])
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORDER
            cell.font = Font(name="Aptos", size=10, color=INK, bold=(col == 2))
            if row % 2:
                cell.fill = PatternFill("solid", fgColor=PORCELAIN)
            apply_value_fill(cell)
        ws.row_dimensions[row].height = 92
    ws.merge_cells("A14:F14")
    ws["A14"] = "WHAT IS ACTUALLY KNOWN VS STILL UNKNOWN"
    ws["A14"].fill = PatternFill("solid", fgColor=ESPRESSO)
    ws["A14"].font = Font(name="Aptos", size=11, bold=True, color=WHITE)
    knowledge = [
        ("Known with high confidence", "Young brands can launch quickly; batch proof and founder visibility matter; creators can create demand shocks; female-leading fronts can share infrastructure; vendors and partners are fragile."),
        ("Known with medium confidence", "Product-name search can be efficient where approved; lifecycle and simple volume ladders can lift economics; long creator relationships outperform one-offs."),
        ("Not known", "True market-wide buyer gender split for direct RUO retatrutide; the anonymous client's settled revenue/margins; Alex's future cohort economics; product-level Biologix net sales without exports."),
        ("Must be measured", "Buyer cohort, creator source, settled contribution, product concentration, repeat interval, refunds, chargebacks, processor reserves, service burden, and referrals."),
    ]
    for row, (label, text) in enumerate(knowledge, 15):
        ws.cell(row, 1, label).font = Font(name="Aptos", bold=True, color=SAGE if "Known" in label else CLAY)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
        ws.cell(row, 2, text)
        ws.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
        for col in range(1, 7):
            ws.cell(row, col).border = BORDER
        ws.row_dimensions[row].height = 58
    ws.freeze_panes = "A4"
    ws.sheet_view.zoomScale = 85


def build_workbook() -> Path:
    wb = Workbook()
    wb.properties.creator = "OpenAI Codex for Alex Weinstein"
    wb.properties.title = "Peptide Founder Media Intelligence"
    wb.properties.subject = "Recent founder/operator media, current verification, and Biologix lessons"
    wb.properties.description = (
        "Dated research workbook separating episode claims from current verification. "
        "No dosing or personal-use instructions; not medical or legal advice."
    )
    wb.calculation.fullCalcOnLoad = True
    build_start_here(wb)
    build_executive_sheet(wb)
    add_tabular_sheet(
        wb,
        "SOURCE LIBRARY",
        SOURCES,
        "All podcast, YouTube, and interview links with company stage, present status, freshness, and confidence.",
        {
            "ID": 11,
            "Priority": 16,
            "Lane": 22,
            "Company": 20,
            "Founder / operator": 27,
            "Episode / source": 38,
            "Published": 13,
            "Duration": 17,
            "Stage at recording": 27,
            "Model": 38,
            "Media link": 14,
            "Current-check link": 17,
            "Current status at 2026-07-26": 52,
            "Freshness": 24,
            "Confidence": 22,
            "Why it matters": 54,
        },
        COBALT,
    )
    add_tabular_sheet(
        wb,
        "TIMESTAMPED NOTES",
        NOTES,
        "Operator detail extracted into decision-ready rows. Timestamp links open the exact YouTube moment where available.",
        {
            "Source ID": 12,
            "Company": 20,
            "Timestamp": 18,
            "Topic": 22,
            "What the operator said / what happened": 55,
            "What Biologix can learn": 55,
            "Evidence class": 33,
            "Freshness note": 32,
            "Timestamp link": 14,
        },
        GOLD,
    )
    add_tabular_sheet(
        wb,
        "CLAIM CHECKS",
        CLAIMS,
        "The most decision-relevant numbers and assertions, with verification status and present-day finding.",
        {
            "Claim ID": 12,
            "Company": 22,
            "Claim to evaluate": 48,
            "Where it came from": 27,
            "Verification status": 26,
            "Current finding at 2026-07-26": 58,
            "Evidence class": 34,
            "Confidence (0-5)": 17,
            "Decision impact": 47,
            "Verification link": 16,
        },
        CLAY,
    )
    ws_claims = wb["CLAIM CHECKS"]
    confidence_col = list(CLAIMS[0].keys()).index("Confidence (0-5)") + 1
    confidence_letter = get_column_letter(confidence_col)
    ws_claims.conditional_formatting.add(
        f"{confidence_letter}6:{confidence_letter}{5 + len(CLAIMS)}",
        CellIsRule(operator="lessThan", formula=["3"], fill=PatternFill("solid", fgColor=RED_LIGHT)),
    )
    ws_claims.conditional_formatting.add(
        f"{confidence_letter}6:{confidence_letter}{5 + len(CLAIMS)}",
        CellIsRule(operator="greaterThanOrEqual", formula=["4"], fill=PatternFill("solid", fgColor=SAGE_LIGHT)),
    )
    add_tabular_sheet(
        wb,
        "CURRENT CONTEXT",
        CURRENT_CONTEXT,
        "What changed recently, what did not change, and when each topic must be checked again.",
        {
            "Date": 24,
            "Topic": 29,
            "What changed": 55,
            "What did NOT change": 55,
            "Business implication": 52,
            "Evidence class": 34,
            "Primary / current link": 16,
            "Next review": 32,
        },
        RED,
    )
    add_tabular_sheet(
        wb,
        "PLAYBOOK",
        LESSONS,
        "Cross-source operating principles, tailored to the female-leading, creator-distributed Biologix thesis.",
        {
            "Domain": 21,
            "Lesson": 48,
            "Evidence": 38,
            "Confidence": 14,
            "How to apply": 64,
            "Priority": 19,
            "Metric / proof": 48,
        },
        SAGE,
    )
    playbook_ws = wb["PLAYBOOK"]
    conf_col = list(LESSONS[0].keys()).index("Confidence") + 1
    conf_letter = get_column_letter(conf_col)
    playbook_ws.conditional_formatting.add(
        f"{conf_letter}6:{conf_letter}{5 + len(LESSONS)}",
        CellIsRule(operator="greaterThanOrEqual", formula=["5"], fill=PatternFill("solid", fgColor=SAGE_LIGHT)),
    )
    add_tabular_sheet(
        wb,
        "ANTI-LESSONS",
        ANTI_LESSONS,
        "Patterns to reject, why they fail, and the safer operating alternative.",
        {
            "Source / pattern": 28,
            "Do not copy": 54,
            "Why": 55,
            "Safer pattern": 60,
            "Risk level": 16,
        },
        CLAY,
    )
    add_tabular_sheet(
        wb,
        "90-DAY ACTION PLAN",
        ACTION_PLAN,
        "A staged, capacity-gated plan that turns the research into measurable decisions without drifting into medical-use guidance.",
        {
            "Window": 18,
            "Workstream": 22,
            "Action": 64,
            "Owner": 24,
            "Dependency": 38,
            "Definition of done": 57,
            "Metric": 45,
            "Priority": 16,
        },
        COBALT,
    )
    add_tabular_sheet(
        wb,
        "WATCHLIST",
        WATCHLIST,
        "What can go stale, the event that should trigger a decision update, and the source to monitor.",
        {
            "Subject": 30,
            "What to watch": 58,
            "Current state": 58,
            "Trigger": 39,
            "Cadence": 15,
            "Owner": 24,
            "Source": 16,
        },
        GOLD,
    )
    # Keep the intended reading order explicit.
    order = [
        "START HERE",
        "EXECUTIVE SYNTHESIS",
        "SOURCE LIBRARY",
        "TIMESTAMPED NOTES",
        "CLAIM CHECKS",
        "CURRENT CONTEXT",
        "PLAYBOOK",
        "ANTI-LESSONS",
        "90-DAY ACTION PLAN",
        "WATCHLIST",
    ]
    wb._sheets = [wb[name] for name in order]
    wb.active = 0
    wb.save(OUTPUT)
    return OUTPUT


def validate_workbook(path: Path) -> None:
    for collection_name, rows in {
        "sources": SOURCES,
        "notes": NOTES,
        "claims": CLAIMS,
        "context": CURRENT_CONTEXT,
        "watchlist": WATCHLIST,
    }.items():
        for idx, row in enumerate(rows, 1):
            for key, value in row.items():
                if ("link" in key.lower() or key == "Source") and value and not valid_url(str(value)):
                    raise ValueError(f"Invalid URL in {collection_name} row {idx} field {key}: {value}")
    check = load_workbook(path, read_only=False, data_only=False)
    expected = [
        "START HERE",
        "EXECUTIVE SYNTHESIS",
        "SOURCE LIBRARY",
        "TIMESTAMPED NOTES",
        "CLAIM CHECKS",
        "CURRENT CONTEXT",
        "PLAYBOOK",
        "ANTI-LESSONS",
        "90-DAY ACTION PLAN",
        "WATCHLIST",
    ]
    if check.sheetnames != expected:
        raise ValueError(f"Unexpected sheet order: {check.sheetnames}")
    if check["SOURCE LIBRARY"].max_row != len(SOURCES) + 5:
        raise ValueError("Source row count mismatch")
    if check["TIMESTAMPED NOTES"].max_row != len(NOTES) + 5:
        raise ValueError("Note row count mismatch")
    if check["CLAIM CHECKS"].max_row != len(CLAIMS) + 5:
        raise ValueError("Claim row count mismatch")
    hyperlink_count = sum(
        1
        for ws in check.worksheets
        for row in ws.iter_rows()
        for cell in row
        if cell.hyperlink is not None
    )
    if hyperlink_count < len(SOURCES) * 2:
        raise ValueError(f"Too few hyperlinks: {hyperlink_count}")
    check.close()
    print(
        f"Built {path}\n"
        f"Sheets: {len(expected)} | sources: {len(SOURCES)} | notes: {len(NOTES)} | "
        f"claims: {len(CLAIMS)} | lessons: {len(LESSONS)} | links: {hyperlink_count}"
    )


if __name__ == "__main__":
    web_output_path = build_web_dataset()
    output_path = build_workbook()
    validate_workbook(output_path)
    print(f"Built {web_output_path}")
